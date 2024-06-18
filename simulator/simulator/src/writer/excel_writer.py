
from data.general import Area, Model, Warnings
from simulation.simulation_lists import *
from util import Tools
from datetime import datetime
from simulation.inventory import Inventory
from scenario import Operation
from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.styles import Alignment, Border, Side, Color, Font
from data.variables import *

import logging
import os
import i18n
import datetime


class ExcelWriter():

    def merge_bold_center(sheet, cell_range, row, col, val):
        """
        Function used to create a bold-merged cell center aligned at the output xlsx file.
        """

        sheet.merge_cells(cell_range)
        cell = sheet.cell(row = row, column = col)
        cell.value = val
        cell.alignment = Alignment(horizontal = 'center')  # this will be modified on plot.py to the plot sheet
        cell.font = cell.font.copy(bold = True)        

    def merge_bold_left(sheet, cell_range, row, col, val):
        """
        Function used to create a bold-merged cell center aligned at the output xlsx file.
        """

        sheet.merge_cells(cell_range)
        cell = sheet.cell(row = row, column = col)
        cell.value = val
        cell.alignment = Alignment(horizontal = 'left')  # this will be modified on plot.py to the plot sheet
        cell.font = cell.font.copy(bold = True)   

    def merge(sheet, cell_range, row, col, val):
        """
        Function used to create a merged cell at the output xlsx file.
        """

        sheet.merge_cells(cell_range)
        cell = sheet.cell(row = row, column = col)
        cell.value = val

    def summary_sheet(self, ws_summary, scenario_name: str, labels: dict, plot, decimals: int = 2):
        """
        Function that print the neccesary information on the summary sheet of the output.
        That function is activated by generate_xslt_file, from simulation.py document.
        """

        ws_summary.title = labels['simanfor.general.summary_sheet']

        # Edit the color of the selected sheet
        ws_summary.sheet_properties.tabColor = "228B22"

        # SIMANFOR logo
        try:
            logo_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "logo.png")
            logo = drawing.image.Image(logo_path)
    
            logo.anchor = 'A1'
            logo.width = 250
            logo.height = 78
            ws_summary.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)

        # Code to stablish the summary cells width
        n = 0  # lines counter
        ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        for k in ascii_uppercase:
            if n < len(ascii_uppercase):
                ws_summary.column_dimensions[ascii_uppercase[n]].width = 12
                n += 1

        # First block of general information
        self.merge_bold_center(ws_summary, 'D1:E1', 1, 4, labels['simanfor.general.study_area'])                       
        self.merge_bold_center(ws_summary, 'D2:E2', 2, 4, labels['simanfor.general.forest'])
        self.merge_bold_center(ws_summary, 'D3:E3', 3, 4, labels['simanfor.general.main_specie'])
        self.merge_bold_center(ws_summary, 'D4:E4', 4, 4, labels['simanfor.general.datetime'])
        self.merge(ws_summary, 'F1:H1', 1, 6, Area.study_area[plot.plot_id])
        self.merge(ws_summary, 'F2:H2', 2, 6, Area.forest[plot.plot_id])
        self.merge(ws_summary, 'F3:H3', 3, 6, Area.main_specie[plot.plot_id])
        self.merge(ws_summary, 'F4:H4', 4, 6, datetime.datetime.now())       
        
        # Second block of general information
        self.merge_bold_center(ws_summary, 'I1:J1', 1, 9, labels['simanfor.general.inventory'])
        self.merge_bold_center(ws_summary, 'I2:J2', 2, 9, labels['simanfor.general.plot'])
        self.merge_bold_center(ws_summary, 'I3:J3', 3, 9, labels['simanfor.general.model'])
        self.merge_bold_center(ws_summary, 'I4:J4', 4, 9, labels['simanfor.general.scenario'])
        self.merge(ws_summary, 'K1:M1', 1, 11, plot.inventory_id)
        self.merge(ws_summary, 'K2:M2', 2, 11, plot.id)
        self.merge(ws_summary, 'K3:M3', 3, 11, Model.model_name)
        self.merge(ws_summary, 'K4:M4', 4, 11, scenario_name)   

        # Blocks of general information - Before cut info
        self.merge_bold_left(ws_summary, 'A6:B6', 6, 1, '')
        self.merge_bold_left(ws_summary, 'C6:F6', 6, 3, labels['simanfor.general.stand_before_cut'])

        if 'AGE' in PLOT_VARS:  # we print at the first summary column the stand age...
            self.merge_bold_center(ws_summary, 'A7:A7', 7, 1, labels['simanfor.general.sum_age'])
        elif 'YEAR' in PLOT_VARS:  # if not, the year of the treatment...
            self.merge_bold_center(ws_summary, 'A7:A7', 7, 1, labels['simanfor.general.sum_year'])
        else:  # if not, the scenario age
            self.merge_bold_center(ws_summary, 'A7:A7', 7, 1, labels['simanfor.plot.scenario_age'])

        self.merge_bold_center(ws_summary, 'B7:B7', 7, 2, labels['simanfor.general.sum_hdom'])
        self.merge_bold_center(ws_summary, 'C7:C7', 7, 3, labels['simanfor.general.sum_density_b_cut'])
        self.merge_bold_center(ws_summary, 'D7:D7', 7, 4, labels['simanfor.general.sum_qmdbh_b_cut'])
        self.merge_bold_center(ws_summary, 'E7:E7', 7, 5, labels['simanfor.general.sum_ba_b_cut'])
        self.merge_bold_center(ws_summary, 'F7:F7', 7, 6, labels['simanfor.general.sum_vol_b_cut'])

        # Blocks of general information - Cut info  
        self.merge_bold_center(ws_summary, 'G6:I6', 6, 7, labels['simanfor.general.stand_cut'])

        self.merge_bold_center(ws_summary, 'G7:G7', 7, 7, labels['simanfor.general.sum_density_cut'])
        self.merge_bold_center(ws_summary, 'H7:H7', 7, 8, labels['simanfor.general.sum_qmdbh_cut'])
        self.merge_bold_center(ws_summary, 'I7:I7', 7, 9, labels['simanfor.general.sum_vol_cut'])

        # Blocks of general information - After cut info        
        self.merge_bold_center(ws_summary, 'J6:M6', 6, 10, labels['simanfor.general.stand_after_cut'])

        self.merge_bold_center(ws_summary, 'J7:J7', 7, 10, labels['simanfor.general.sum_density_a_cut'])
        self.merge_bold_center(ws_summary, 'K7:K7', 7, 11, labels['simanfor.general.sum_qmdbh_a_cut'])
        self.merge_bold_center(ws_summary, 'L7:L7', 7, 12, labels['simanfor.general.sum_ba_a_cut'])
        self.merge_bold_center(ws_summary, 'M7:M7', 7, 13, labels['simanfor.general.sum_vol_a_cut'])

        # Information about dead, ingrowth and special data will be printed only if it is calculated at the model --> plot.py


    def description_sheet(self, ws_description, scenario_name: str, labels: dict, plot, decimals: int = 2):
        """
        Function that print the neccesary information on the description sheet of the output.
        That function is activated by generate_xslt_file, from simulation.py document.
        """

        # Edit the color of the selected sheet
        ws_description.sheet_properties.tabColor = "996D14"

        # logo SIMANFOR
        try:
            logo_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "logo.png")
            logo = drawing.image.Image(logo_path)
    
            logo.anchor = 'A1'
            logo.width = 250
            logo.height = 80
            ws_description.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)

        # logo iuFOR
        try:
            logo_iufor = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "iufor.png")
            logo = drawing.image.Image(logo_iufor)
    
            logo.anchor = 'J1'
            logo.width = 250
            logo.height = 80
            ws_description.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)

        # Edit the color of the cell borders
        for k in range(1,4):  # web simanfor
            for j in range(5,9):
                ws_description.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '613605'))
        for k in range(1,4):  # web iufor
            for j in range(14,18):
                ws_description.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '03B300'))
        for k in range(5,7):  # titles
            for j in range(1,21):
                ws_description.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '6DD505'))      
        for k in range(10,12):  # titles
            for j in range(10,21):
                ws_description.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '6DD505'))      


        # SIMANFOR and iuFOR links
        self.merge_bold_center(ws_description, 'E2:H2', 2, 5, labels['simanfor.metadata.web_sima'])
        self.merge_bold_center(ws_description, 'E3:H3', 3, 5, labels['simanfor.metadata.link_sima'])
        self.merge_bold_center(ws_description, 'N2:Q2', 2, 14, labels['simanfor.metadata.web_iufor'])
        self.merge_bold_center(ws_description, 'N3:Q3', 3, 14, labels['simanfor.metadata.link_iufor'])

        # Information - Study area - titles and contents

        lista_area_vars = (  # copied from from_plot_to_area from plot.py file
        Area.plot_type[plot.plot_id],
        Area.plot_area[plot.plot_id],
        Area.province[plot.plot_id],
        Area.study_area[plot.plot_id],
        Area.municipality[plot.plot_id],
        Area.forest[plot.plot_id],
        Area.prov_region[plot.plot_id],
        Area.main_specie[plot.plot_id],
        Area.specie_ifn_id[plot.plot_id],
        Area.slope[plot.plot_id],
        Area.aspect[plot.plot_id],
        Area.continentality[plot.plot_id],
        Area.altitude[plot.plot_id],
        Area.longitude[plot.plot_id],
        Area.latitude[plot.plot_id],   
        Area.aa_rainfall[plot.plot_id],
        Area.ma_temperature[plot.plot_id],
        Area.september_rain[plot.plot_id],
        Area.september_temp[plot.plot_id],
        Area.november_rain[plot.plot_id],
        Area.november_temp[plot.plot_id],
        Area.martonne[plot.plot_id],
        Area.martonne_2020[plot.plot_id],
        Area.martonne_2040[plot.plot_id],
        Area.martonne_2060[plot.plot_id],                
        Area.martonne_2080[plot.plot_id],   
        # Cistus ladanifer special variables    
        Area.tr[plot.plot_id],
        Area.rain_as[plot.plot_id],
        Area.tmin_so[plot.plot_id],
        Area.tmin_on[plot.plot_id],
        Area.tmin_ond[plot.plot_id],
        Area.tmmin_oct[plot.plot_id],
        Area.tsum_mean_so[plot.plot_id],
        Area.tsum_mmin_so[plot.plot_id],
        Area.tsum_mmin_on[plot.plot_id],
        Area.tsum_mmin_sond[plot.plot_id])

        row_area = 7  # starting row
        
        self.merge_bold_center(ws_description, 'A6:I6', 6, 1, labels['simanfor.general.study_area']) 
        
        for k in range(len(AREA_VARS)):

            if lista_area_vars[k] != '':
                self.merge_bold_center(ws_description, 'A' + str(row_area) + ':C' + str(row_area), row_area, 1, labels['simanfor.area.' + AREA_VARS[k]])
                self.merge(ws_description, 'D' + str(row_area) + ':I' + str(row_area), row_area, 4, lista_area_vars[k]) 
                row_area += 1


        # Information - Plot information - titles
        self.merge_bold_center(ws_description, 'J6:T6', 6, 10, labels['simanfor.general.plot_info']) 
        if 'REINEKE_VALUE' in PLOT_VARS:
            self.merge_bold_center(ws_description, 'J7:L7', 7, 10, labels['simanfor.plot.REINEKE_VALUE']) 
        if 'REF_SI_AGE' in PLOT_VARS:
            self.merge_bold_center(ws_description, 'J8:L8', 8, 10, labels['simanfor.plot.REF_SI_AGE']) 
        if 'SI' in PLOT_VARS:   
            self.merge_bold_center(ws_description, 'J9:L9', 9, 10, labels['simanfor.plot.SI']) 
        #if 'HEGYI_RADIUS' in PLOT_VARS:   
        #    self.merge_bold_center(ws_description, 'J10:L19', 10, 10, labels['simanfor.plot.HEGYI_RADIUS']) 

        # Information - Model information - titles
        self.merge_bold_center(ws_description, 'J11:T11', 11, 10, labels['simanfor.general.model_info'])  
        self.merge_bold_center(ws_description, 'J12:L12', 12, 10, labels['simanfor.model.MODEL_NAME']) 
        self.merge_bold_center(ws_description, 'J13:L13', 13, 10, labels['simanfor.model.SPECIE_IFN_ID']) 
        self.merge_bold_center(ws_description, 'J14:L14', 14, 10, labels['simanfor.model.APLICATION_AREA']) 
        self.merge_bold_center(ws_description, 'J15:L15', 15, 10, labels['simanfor.model.VALID_PROV_REG']) 
        self.merge_bold_center(ws_description, 'J16:L16', 16, 10, labels['simanfor.model.EXEC_TIME']) 
        self.merge_bold_center(ws_description, 'J17:L17', 17, 10, labels['simanfor.model.MODEL_TYPE']) 
        self.merge_bold_center(ws_description, 'J18:L18', 18, 10, labels['simanfor.model.MODEL_CARD_ES']) 
        self.merge_bold_center(ws_description, 'J19:L19', 19, 10, labels['simanfor.model.MODEL_CARD_EN']) 
                
        # Information - Plot information - titles
        #self.merge(ws_description, 'M11:T11', 11, 13, str(plot.reineke_value)) 
        #self.merge(ws_description, 'M12:T12', 12, 13, str(plot.si))           # THAT VARIABLES ARE PRINTED ON plot_to_xls of plot.py
        #self.merge(ws_description, 'M13:T13', 13, 13, str(plot.ref_si_age))   # because from here data content is not available yet 

        # Information - Model information - contents
        self.merge(ws_description, 'M12:T12', 12, 13, Model.model_name) 
        self.merge(ws_description, 'M13:T13', 13, 13, (Model.specie_ifn_id)) 
        self.merge(ws_description, 'M14:T14', 14, 13, Model.aplication_area) 
        self.merge(ws_description, 'M15:T15', 15, 13, (Model.valid_prov_reg)) 
        self.merge(ws_description, 'M16:T16', 16, 13, (Model.exec_time)) 
        self.merge(ws_description, 'M17:T17', 17, 13, labels['simanfor.metadata.' + Model.model_type]) 
        self.merge(ws_description, 'M18:T18', 18, 13, Model.model_card_es) 
        self.merge(ws_description, 'M19:T19', 19, 13, Model.model_card_en) 


    def metadata_sheet(self, ws_metadata, scenario_name: str, labels: dict, plot, decimals: int = 2):
        """
        Function that print the neccesary information on the metadata sheet of the output.
        That function is activated by generate_xslt_file, from simulation.py document.
        """

        # Edit the color of the selected sheet
        ws_metadata.sheet_properties.tabColor = "613605"

        # logo SIMANFOR
        try:
            logo_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "logo.png")
            logo = drawing.image.Image(logo_path)
    
            logo.anchor = 'A1'
            logo.width = 250
            logo.height = 80
            ws_metadata.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)

        # logo iuFOR
        try:
            logo_iufor = os.path.join(os.path.dirname(os.path.realpath(__file__)), "..", "files", "iufor.png")
            logo = drawing.image.Image(logo_iufor)
    
            logo.anchor = 'J1'
            logo.width = 250
            logo.height = 80
            ws_metadata.add_image(logo)
        except:
            Tools.print_log_line("Couldn't find logo image.", logging.INFO)


        # We add "special information" of the summary sheet, only in the case that they are present at the model
        if 'DEAD_DENSITY' in PLOT_VARS:  # if we have dead trees to eliminate from the plot...
            SUMMARY.append('stand_dead')
        
        if 'ING_DENSITY' in PLOT_VARS:  # if we have trees to add at the plot...
            SUMMARY.append('stand_ingrowth')

        if 'W_CORK' in PLOT_VARS: # Qsuber variables
            SUMMARY.append('QSUBER_VARS')
            SUMMARY.append('W_CORK')
            SUMMARY.append('BARK_VOL')

        if 'ALL_CONES' in PLOT_VARS:  # Ppinea variables
            SUMMARY.append('PPINEA_VARS') 
            SUMMARY.append('ALL_CONES') 
            SUMMARY.append('SOUND_CONES') 
            SUMMARY.append('SOUND_SEEDS') 
            SUMMARY.append('W_SOUND_CONES') 
            SUMMARY.append('W_ALL_CONES') 

        if 'EDIBLE_MUSH' in PLOT_VARS:  # Mushroom variables
            SUMMARY.append('MUSHROOMS_VARS') 
            SUMMARY.append('EDIBLE_MUSH') 
            SUMMARY.append('MARKETED_MUSH') 
            SUMMARY.append('MARKETED_LACTARIUS') 

        if 'scenario_id' in SCENARIO_VARS:
            SCENARIO_VARS.remove('scenario_id')  # that variable is not needed at the output

        # Before print too, I calculate the longest lenght of each pair of lists
        if len(SUMMARY) > len(AREA_VARS):
            len_1 = len(SUMMARY)
        else:
            len_1 = len(AREA_VARS)
        if len(MODEL_VARS) > len(SCENARIO_VARS):
            len_2 = len(MODEL_VARS)
        else:
            len_2 = len(SCENARIO_VARS)

        # Edit the color of the cell borders
        for k in range(1,4):  # web simanfor
            for j in range(5,9):
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '613605'))
        for k in range(1,4):  # web iufor
            for j in range(14,18):
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '03B300'))
        for k in range(5,9):  # models
            for j in range(1,7):
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = 'A15908'))
        #for k in range(5,9):  # webs
        #    for j in range(8,21):
        #        ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '2DD51A'))
        for k in range(9,12):  # vars historial, summary and study area vars
            for j in range(1,21):
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '11288D'))
        for k in range(13 + len_1, 15 + len_1):
            for j in range(1,21):  # plot and scenario vars
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '11288D'))
        for k in range(16 + len_1 + len_2, 18 + len_1 + len_2):
            for j in range(1,21):  # cuts vars
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '11288D'))
        for k in range(20 + len_1 + len_2 + 3, 22 + len_1 + len_2 + 3): 
            for j in range(1,21):  # plot vars
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '11288D'))
        for k in range(23 + len_1 + len_2 + 3 + len(PLOT_VARS), 25 + len_1 + len_2 + 3 + len(PLOT_VARS)): 
            for j in range(1,21):  # tree vars
                ws_metadata.cell(row=k, column=j).border = Border(bottom = Side(border_style = 'thick', color = '11288D'))

        # general information and links
        self.merge_bold_center(ws_metadata, 'E2:H2', 2, 5, labels['simanfor.metadata.web_sima'])
        self.merge_bold_center(ws_metadata, 'E3:H3', 3, 5, labels['simanfor.metadata.link_sima'])
        self.merge_bold_center(ws_metadata, 'N2:Q2', 2, 14, labels['simanfor.metadata.web_iufor'])
        self.merge_bold_center(ws_metadata, 'N3:Q3', 3, 14, labels['simanfor.metadata.link_iufor'])

        #self.merge_bold_center(ws_metadata, 'H6:H6', 6, 8, labels['simanfor.metadata.facebook'])
        #self.merge_bold_center(ws_metadata, 'I6:M6', 6, 9, labels['simanfor.metadata.iufor_facebook'])
        #self.merge_bold_center(ws_metadata, 'H7:H7', 7, 8, labels['simanfor.metadata.twitter'])
        #self.merge_bold_center(ws_metadata, 'I7:M7', 7, 9, labels['simanfor.metadata.iufor_twitter'])
        #self.merge_bold_center(ws_metadata, 'H8:H8', 8, 8, labels['simanfor.metadata.instagram'])
        #self.merge_bold_center(ws_metadata, 'I8:M8', 8, 9, labels['simanfor.metadata.iufor_instagram'])
        #self.merge_bold_center(ws_metadata, 'N6:N6', 6, 14, labels['simanfor.metadata.linkedin'])
        #self.merge_bold_center(ws_metadata, 'O6:T6', 6, 15, labels['simanfor.metadata.iufor_linkedin'])
        #self.merge_bold_center(ws_metadata, 'N7:N7', 7, 14, labels['simanfor.metadata.youtube'])
        #self.merge_bold_center(ws_metadata, 'O7:T7', 7, 15, labels['simanfor.metadata.iufor_youtube'])
        #self.merge_bold_center(ws_metadata, 'N8:N8', 8, 14, labels['simanfor.metadata.flickr'])
        #self.merge_bold_center(ws_metadata, 'O8:T8', 8, 15, labels['simanfor.metadata.iufor_flickr'])

        self.merge_bold_center(ws_metadata, 'A6:F6', 6, 1, labels['simanfor.metadata.model'])
        self.merge_bold_center(ws_metadata, 'A7:F7', 7, 1, Model.model_name)
        self.merge_bold_center(ws_metadata, 'A8:F8', 8, 1, labels['simanfor.metadata.' + Model.model_type]) 

        self.merge_bold_center(ws_metadata, 'A10:T10', 10, 1, labels['simanfor.metadata.vars'])
        self.merge_bold_center(ws_metadata, 'A11:I11', 11, 1, labels['simanfor.metadata.summary'])
        self.merge_bold_center(ws_metadata, 'J11:T11', 11, 10, labels['simanfor.metadata.area'])
        self.merge_bold_center(ws_metadata, 'A' + str(14 + len_1) + ':I' + str(14 + len_1), 14 + len_1, 1, labels['simanfor.metadata.plot_type'])
        self.merge_bold_center(ws_metadata, 'J' + str(14 + len_1) + ':T' + str(14 + len_1), 14 + len_1, 10, labels['simanfor.metadata.scenario'])
        self.merge_bold_center(ws_metadata, 'A' + str(17 + len_1 + len_2) + ':T' + str(17 + len_1 + len_2), 17 + len_1 + len_2, 1, labels['simanfor.metadata.cuts'])
        self.merge_bold_center(ws_metadata, 'A' + str(21 + len_1 + len_2 + 3) + ':T' + str(21 + len_1 + len_2 + 3), 21 + len_1 + len_2 + 3, 1, labels['simanfor.metadata.plot'])
        self.merge_bold_center(ws_metadata, 'A' + str(24 + len_1 + len_2 + 3 + len(PLOT_VARS)) + ':T' + str(24 + len_1 + len_2 + 3 + len(PLOT_VARS)), 24 + len_1 + len_2 + 3 + len(PLOT_VARS), 1, labels['simanfor.metadata.tree'])

        # print the variable names and their explanations
        for j in range(len(SUMMARY)):  # summary
            self.merge_bold_center(ws_metadata, ('A' + str(13 + j) + ':C' + str(13 + j)), 13 + j, 1, labels['simanfor.general.' + SUMMARY[j]])
            self.merge(ws_metadata, ('D' + str(13 + j) + ':I' + str(13 + j)), 13 + j, 4, labels['simanfor.metadata.' + SUMMARY[j]])
        for j in range(len(AREA_VARS)):  # area
                self.merge_bold_center(ws_metadata, ('J' + str(13 + j) + ':L' + str(13 + j)), 13 + j, 10, labels['simanfor.area.' + AREA_VARS[j]])
                self.merge(ws_metadata, ('M' + str(13 + j) + ':T' + str(13 + j)), 13 + j, 13, labels['simanfor.metadata.' + AREA_VARS[j]]) 
        for j in range(len(MODEL_VARS)):  # model
            self.merge_bold_center(ws_metadata, ('A' + str(16 + j + len_1) + ':C' + str(16 + j + len_1)), 16 + j + len_1, 1, labels['simanfor.model.' + MODEL_VARS[j]])
            self.merge(ws_metadata, ('D' + str(16 + j + len_1) + ':I' + str(16 + j + len_1)), 16 + j + len_1, 4, labels['simanfor.metadata.' + MODEL_VARS[j]])
        for j in range(len(SCENARIO_VARS)):  # scenario
            self.merge_bold_center(ws_metadata, ('J' + str(16 + j + len_1) + ':L' + str(16 + j + len_1)), 16 + j + len_1, 10, labels['simanfor.plot.' + SCENARIO_VARS[j]])
            self.merge(ws_metadata, ('M' + str(16 + j + len_1) + ':T' + str(16 + j + len_1)), 16 + j + len_1, 13, labels['simanfor.metadata.' + SCENARIO_VARS[j]])  
        for j in range(len(CUTS)):  # cuts
            if j < 3:
                self.merge_bold_center(ws_metadata, ('A' + str(20 + j + len_1 + len_2) + ':C' + str(20 + j + len_1 + len_2)), 20 + j + len_1 + len_2, 1, labels['simanfor.general.' + CUTS[j]])
                self.merge(ws_metadata, ('D' + str(20 + j + len_1 + len_2) + ':I' + str(20 + j + len_1 + len_2)), 20 + j + len_1 + len_2, 4, labels['simanfor.metadata.' + CUTS[j]])     
            else:
                self.merge_bold_center(ws_metadata, ('J' + str(17 + j + len_1 + len_2) + ':L' + str(17 + j + len_1 + len_2)), 17 + j + len_1 + len_2, 10, labels['simanfor.general.' + CUTS[j]])
                self.merge(ws_metadata, ('M' + str(17 + j + len_1 + len_2) + ':T' + str(17 + j + len_1 + len_2)), 17 + j + len_1 + len_2, 13, labels['simanfor.metadata.' + CUTS[j]])     
        # cut types and criterias
        self.merge_bold_center(ws_metadata, ('A' + str(19 + len_1 + len_2) + ':C' + str(19 + len_1 + len_2)), 19 + len_1 + len_2, 1, labels['simanfor.plot.cut_type'])
        self.merge(ws_metadata, ('D' + str(19 + len_1 + len_2) + ':I' + str(19 + len_1 + len_2)), 19 + len_1 + len_2, 4, labels['simanfor.plot.cut_type'])   
        self.merge_bold_center(ws_metadata, ('J' + str(19 + len_1 + len_2) + ':L' + str(19 + len_1 + len_2)), 19 + len_1 + len_2, 10, labels['simanfor.plot.cut_criteria'])  
        self.merge(ws_metadata, ('M' + str(19 + len_1 + len_2) + ':T' + str(19 + len_1 + len_2)), 19 + len_1 + len_2, 13, labels['simanfor.plot.cut_criteria'])                   
        for j in range(len(PLOT_VARS)):  # plot
            self.merge_bold_center(ws_metadata, ('A' + str(23 + j + len_1 + len_2 + 3) + ':C' + str(23 + j + len_1 + len_2 + 3)), 23 + j + len_1 + len_2 + 3, 1, labels['simanfor.plot.' + PLOT_VARS[j]])      
            self.merge(ws_metadata, ('D' + str(26 + j + len_1 + len_2) + ':T' + str(26 + j + len_1 + len_2)), 26 + j + len_1 + len_2, 4, labels['simanfor.metadata.' + PLOT_VARS[j]])     
        for j in range(len(TREE_VARS)):  # tree
            self.merge_bold_center(ws_metadata, ('A' + str(27 + j + len_1 + len_2 + len(PLOT_VARS) + 3) + ':C' + str(27 + j + len_1 + len_2 + len(PLOT_VARS) + 3)), 27 + j + len_1 + len_2 + len(PLOT_VARS) + 3, 1, labels['simanfor.tree.' + TREE_VARS[j]])
            self.merge(ws_metadata, ('D' + str(30 + j + len_1 + len_2 + len(PLOT_VARS)) + ':T' + str(30 + j + len_1 + len_2 + len(PLOT_VARS))), 30 + j + len_1 + len_2 + len(PLOT_VARS), 4, labels['simanfor.metadata.' + TREE_VARS[j]])              
        self.merge_bold_center(ws_metadata, ('A' + str(26 + len_1 + len_2 + len(PLOT_VARS) + 3) + ':C' + str(26 + len_1 + len_2 + len(PLOT_VARS) + 3)), 26 + len_1 + len_2 + len(PLOT_VARS) + 3, 1, labels['simanfor.tree.status'])
        self.merge(ws_metadata, ('D' + str(29 + len_1 + len_2 + len(PLOT_VARS)) + ':T' + str(29 + len_1 + len_2 + len(PLOT_VARS))), 29 + len_1 + len_2 + len(PLOT_VARS), 4, labels['simanfor.metadata.status'])  
