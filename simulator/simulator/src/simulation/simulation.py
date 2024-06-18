#!/usr/bin/env python
#
# Copyright (c) $today.year Moises Martinez (Sngular). All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

from data import Tree
from data import Plot
from data.general import Area, Model, Warnings
from scenario.step import Step
from .simulation_lists import *
from util import Tools
from datetime import datetime
from .inventory import Inventory
from scenario import Operation
from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.styles import Alignment, Border, Side, Color, Font, PatternFill
from models import HarvestModel
from data.variables import *
from writer.excel_writer import ExcelWriter
from writer.labels import Labels

import logging
import os
import i18n
import json

from dask import delayed

XLSX = 0
JSON = 1

class Simulation:

    def __init__(self, date=datetime.now()):

        self.__date = date
        self.__steps = list()

    def add_step(self, step_id, inventory: Inventory, operation: Operation, model):

        min_age = 0
        max_age = 0

        if len(self.__steps) == 0:
            age = operation.get_variable('init')
        else:
            last = self.get_last()
            if operation.get_variable('time') == None:
                age = last.age
            else:
                age = last.age + operation.get_variable('time')
            min_age = operation.get_variable('min_age')
            max_age = operation.get_variable('max_age')

        model_name = ''

        if isinstance(model, HarvestModel):
            model_name = i18n.t('simanfor.general.' + model.name)

        self.__steps.append(Step(step_id, inventory, operation.type, operation.description, 
                                 age, min_age, max_age, operation, model_name))

    def get_step(self, position):
        if position < len(self.__steps):
            return self.__steps[position]
        return None

    def get_first_step(self):
        return self.get_step(0)

    def get_last(self):
        if len(self.__steps) == 0:
            return 0
        return self.__steps[len(self.__steps)-1]


    def generate_json_file(self, scenario_name: str, file_path, plot_id: str):

        # plots = self.__steps[0].get_plot_ids()

        general = dict()

        summary = dict()
        steps = dict()

        summary[i18n.t('simanfor.general.main_specie')] = ''
        summary[i18n.t('simanfor.general.main_specie')] = ''
        summary[i18n.t('simanfor.general.forest')] = ''
        summary[i18n.t('simanfor.general.study_area')] = ''
        summary[i18n.t('simanfor.general.model')] = ''
        summary[i18n.t('simanfor.general.scenario')] = scenario_name
        summary[i18n.t('simanfor.general.inventory')] = ''
        #summary[i18n.t('simanfor.general.template')] = ''
        summary[i18n.t('simanfor.general.execution')] = ''

        row = 1

        for step in self.__steps:
            steps[row] = step.to_json(plot_id, SCENARIO_VARS, row)
            row += 1

        general[i18n.t('simanfor.general.' + DEFAULT_EXCEL_OUTPUT_FILE_STRUCTURE[0])] = summary
        general[i18n.t('simanfor.general.' + DEFAULT_EXCEL_OUTPUT_FILE_STRUCTURE[1])] = steps

        with open(file_path, 'w') as outfile:
            json.dump(general, outfile)


    def plot_sheet(self, ws_plot, workbook, scenario_name: str, scenario_file_name:str, labels: dict, plot, decimals: int = 2):
        """
        Function that print the neccesary information on the plot sheet of the output.
        That function is activated by generate_xslt_file, from that document.
        """

        # Edit the color of the selected sheet
        ws_plot.sheet_properties.tabColor = "0C5C00"
        
        # Edit group of variables color
        scenarioFill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        plotFill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

        n = 0  # column counter
        z = 2  # number to multiply the lenght of the list        
        ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        for j in range(len(SCENARIO_VARS)):  # exclude file_name from the list to not show the scenario file name at the output file
            ws_plot.cell(row = 1, column = j + 1).value = labels['simanfor.plot.' + SCENARIO_VARS[j]]         

            ws_plot[ascii_uppercase[n] + str(1)].fill = scenarioFill  # set scenario color
            n += 1

        for j in range(len(PLOT_VARS)):
            if PLOT_VARS[j] not in PLOT_VARS_NOT_PRINT:
                ws_plot.cell(row = 1, column = j + 1 + len(SCENARIO_VARS)).value = labels['simanfor.plot.' + PLOT_VARS[j]]        

                if n < len(ascii_uppercase):  # set plot color
                    ws_plot[ascii_uppercase[n] + str(1)].fill = plotFill
                    n += 1
                else:  # once alphabet is finished, we add another first letter
                    if n < len(ascii_uppercase)*z:
                        ws_plot[ascii_uppercase[z - 2] + ascii_uppercase[n - len(ascii_uppercase)*(z - 1)] + str(1)].fill = plotFill
                        n += 1
                    else:
                        z += 1
                        ws_plot[ascii_uppercase[z - 2] + ascii_uppercase[n - len(ascii_uppercase)*(z - 1)] + str(1)].fill = plotFill
                        n += 1


        step_count = 1
        row = 8
        summary_row = 8

        for step in self.__steps:

            next_step = None if step_count >= len(self.__steps) else self.__steps[step_count]

            next_operation = None if next_step == None else next_step._Step__full_operation          
            summary_row = step.to_xslt(labels, workbook, plot.id, scenario_file_name, row, next_step, next_operation, 
                                       summary_row, decimals)
            row += 1
            step_count += 1


    def generate_xslt_file(self, scenario_name: str, scenario_file_name:str, labels: dict, file_path: str, plot, modelo: str, zip_compression: bool = False, 
                           type: int = JSON, decimals: int = 2):
        """
        Function used to create a xlsx output file.
        Summary, metadata and plot sheets are modified by using that function.
        Variables to print must be created at the next function, named get_labels.
        """

        Tools.print_log_line('Generating xslt file for plot ' + str(plot.id), logging.INFO)

        workbook = Workbook()  # create an excel workbook
        ws_summary = workbook.active  # activate it, and create the first sheet
        
        # cover the first sheet with the summary info
        ExcelWriter.summary_sheet(ExcelWriter, ws_summary, scenario_name, labels, plot, decimals)  

        # create and cover the second sheet with the description info
        ws_description = workbook.create_sheet(labels['simanfor.general.description_sheet'])
        ExcelWriter.description_sheet(ExcelWriter, ws_description, scenario_name, labels, plot, decimals)  

        # create and cover the third sheet with the metadata info
        ws_metadata = workbook.create_sheet(labels['simanfor.general.metadata_sheet'])
        ExcelWriter.metadata_sheet(ExcelWriter, ws_metadata, scenario_name, labels, plot, decimals)  

        # create and cover the fourth sheet with the plot info
        ws_plot = workbook.create_sheet(labels['simanfor.general.plot_sheet'])
        self.plot_sheet(ws_plot, workbook, scenario_name, scenario_file_name, labels, plot, decimals)  

        # close the excel file and save it
        workbook.close()
        workbook.save(file_path)
        
        return plot


    def generate_results(self, scenario_name: str, scenario_file_name: str, file_path: str, modelo: str, type: int = XLSX, zip_compression: bool = False, 
                         decimals: int = 2):

        labels = Labels.get_labels()

        plots = self.get_first_step().inventory.plots

        for plot in plots:

            if isinstance(plot.id, str): # json input
                filename = file_path + OUTPUT_FILE_BASE + os.path.split(plot.id)[1] + '.' + OUTPUT_EXTENSION[type]
            else:
                filename = file_path + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type]

            if type == XLSX:

                print('PRINTING PLOT', plot.id, 'OF INVENTORY', plot.inventory_id,'ON THE OUTPUT...')

                self.generate_xslt_file(
                    scenario_name,
                    scenario_file_name,
                    labels,
                    filename,
                    plot, modelo, decimals=decimals)

            # else:
            #     self.generate_json_file(
            #         name,
            #         current_folder + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type],
            #         plot.id)


    def print_plots(self, plots: list):
        
        # count = 0
        # for plot in plots:
        #     count += 1
        #     # self.print_plot(plot)
            
        return len(plots)


    def generate_results_parallel(self, scenario_name: str, scenario_file_name:str, file_path: str, modelo: str, type: int = XLSX, zip_compression: bool = False, 
                         decimals: int = 2):

        labels = Labels.get_labels()
        # plot_labels['simanfor.general.Summary'] = i18n.t('simanfor.general.Summary')

        plots = self.get_first_step().inventory.plots

        outfile_list = []

        for plot in plots:
            if type == XLSX:
                # recalc_list.append(self.recalculate_process(new_plot, model, 
                #         operation.get_variable('time'), result_pies_mayores, result_inventory))

                outfile_list.append(delayed(self.generate_xslt_file)(scenario_name, scenario_file_name, labels,
                    file_path + OUTPUT_FILE_BASE + str(plot.id) + '.' + OUTPUT_EXTENSION[type], 
                    plot, modelo, decimals=decimals))

        # result_inventory = (delayed(result_inventory.add_plots)(recalc_list)).compute(scheduler="distributed") # default, same as empty ""
        nbr_plots = delayed(self.print_plots)(outfile_list).compute(scheduler="distributed")
        # nbr_plots = delayed(self.print_plots)(outfile_list).compute(scheduler="single-threaded")
        print("Printed", nbr_plots, "output files.")
