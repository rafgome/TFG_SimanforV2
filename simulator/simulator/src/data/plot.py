#!/usr/bin/env python3
#
# Copyright (c) $today.year Moises Martinez (Sngular). All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================

from .tree import Tree
from .general import Area, Model, Warnings
from util import Tools
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Color, PatternFill, Font
from .search.order_criteria import OrderCriteria
from data.variables import *  
from data.inventory_translations.es import ES_PLOT
from data.inventory_translations.gl import GL_PLOT
from data.inventory_translations.en import EN_PLOT

import numpy as np
import logging
import math
import i18n

# from time import sleep

ID = 'PLOT_ID'

DESC = 1
ASC = 2


class Plot:

    @staticmethod
    def get_index_by_name(name):
        return PLOT_VARS.index(name)

    @staticmethod
    def get_name_by_index(index):
        return PLOT_VARS[index]

    def __init__(self, data=None):

        self.__values = dict()  # dictionary useful to set the values for each variable
        self.__trees = dict()  # dicionary useful to save the information about alive trees
        self.__dead_trees = dict()  # dictionary useful to basic_engine calculations about dead trees
        self.__cut_trees = dict()  # dictionary useful to basic_engine calculations about cut trees
        self.__ingrowth_trees = dict()  # dictionary useful to basic_engine calculations about ingrowth trees

        if data is None:  # at the first time, when variables are uploaded to the simulator...
            Tools.print_log_line("No data info. The Plot has been created empty.", logging.WARNING)
            for variable in PLOT_VARS:
                self.__values[variable] = 0  # the value of the variables is 0
        else:  # once the information of initial inventory is uploaded...
            for variable in PLOT_VARS:
                if variable not in data.keys():  # if the variable are not at the initial inventory...
                    #Tools.print_log_line(str(variable) + ' is not in data document', logging.WARNING) 
                    self.__values[variable] = ''  # we set their value as empty
                    self.translate_name(data, variable) 
                else:  # if it is at the initial inventory...
                    self.__values[variable] = data[variable]  # we upload their value

            #self.__values[ID] = int(self.__values[ID])
            self.__values[ID] = self.__values[ID]

            if 'plot' in data: # json input
                self.map_json_to_xl(data)
                
    def translate_name(self, data, variable):
        """
        Function created to translate the names of initial inventory to the variables names of SIMANFOR.
        It is activated from __init__ function, and it uses the lists of variables from inventory_translations.
        """

        lang_list = (ES_PLOT, EN_PLOT, GL_PLOT)  # list of available languages

        for lang in lang_list:  # for each language...
            for variable_lang in lang: # for each variable on the language...
                if variable == variable_lang:  # if the variable is on the language

                    try:  # condition needed to skip non used languages

                        self.__values[variable] = data[lang[variable_lang]]  # their value is added to the SIMANFOR variable

                    except KeyError:

                        None


    def map_json_to_xl(self, data):
        """
        That function converts the name of the json input file variables to the simulator variables names.
        """

        self.__values['PLOT_ID'] = data['plot']
        self.__values['PROVINCE'] = data['provincia']
        self.__values['LATITUDE'] = data['plotlat']
        self.__values['LONGITUDE'] = data['plotlong']
        self.__values['AGE'] = data['age']
        

    def get_value(self, var: str):
        """
        Function neccesary to obtain the variable value
        """
        return self.__values[var]

    def add_tree(self, tree: Tree):
        """
        That function includes new trees to the output file.
        It is activated by basic_engine file, using apply_tree_model function.
        """

        if tree.get_value('status') is None:  # No status indicates that the tree is alive
            self.__trees[tree.id] = tree
        elif tree.get_value('status') == 'M':  # M status indicates that the tree is dead
            self.__dead_trees[tree.id] = tree
        elif tree.get_value('status') == 'C':  # C status indicates that the tree is cut
            self.__cut_trees[tree.id] = tree
        elif tree.get_value('status') == 'I':  # I status indicates that the tree is new (incorporated to the stand)
            self.__ingrowth_trees[tree.id] = tree

    def add_trees(self, trees: list):
        """
        That function includes new trees to the output file.
        It is activated by basic_engine file, using apply_tree_model function.
        """

        for tree in trees:

            if tree.get_value('status') is None or tree.get_value('status') == '':  # No status indicates that the tree is alive
                self.__trees[tree.id] = tree
            elif tree.get_value('status') == 'M':  # M status indicates that the tree is dead
                self.__dead_trees[tree.id] = tree
            elif tree.get_value('status') == 'C':  # C status indicates that the tree is cut
                self.__cut_trees[tree.id] = tree
            elif tree.get_value('status') == 'I':  # I status indicates that the tree is new (incorporated to the stand)
                self.__ingrowth_trees[tree.id] = tree


    def get_tree(self, id: int):
        return self.__trees[id] if id in self.__trees.keys() else None

    def get_tree_by_position(self, position: int):
        count = 0
        for tree in self.__trees:
            if count == position:
                return tree
            count += 1
        return None

    def get_trees_order_by(self, variable, value):
        return sorted(self.__trees, key=lambda x: x.get_values(variable) <= value, reverse=True)

    def get_first_tree_order_by(self, variable, value):
        return None

    @property
    def id(self):
        return self.__values[ID] if ID in self.__values.keys() else None

    @property
    def trees(self):
        return self.__trees.values()

    @property
    def values(self):
        return self.__values

    def get_number_trees(self):
        return len(self.__trees)

    def get_trees_array(self):
        tmp = list()
        for tree in self.__trees.values():
            tmp.append(tree.get_array())
        return tmp

    def get_np_trees_array(self):
        tmp = np.array([])
        for tree in self.__trees.values():
            tmp.append(tree.get_array())
        return tmp

    def short_trees_on_list(self, variable: str, order: int = DESC):
        if order == DESC:
            return sorted(self.__trees.values(), key=lambda tree: tree.get_value(variable), reverse=True)
        else:
            return sorted(self.__trees.values(), key=lambda tree: tree.get_value(variable), reverse=False)

    def add_value(self, variable, value):
        """
        Function neccesary to add a value to a plot variable by sustitution from the past value.
        """
        self.__values[variable] = value

    def set_value(self, variable, value):
        """
        Function neccesary to add a value to a plot variable by sustitution from the past value. Is the same as add_value.
        """
        self.__values[variable] = value

    def sum_value(self, variable, value):
        """
        Function neccesary to sum a value to a plot variable.
        """
        self.__values[variable] += value

    def sub_value(self, variable, value):
        """
        Function neccesary to subtrack a value to a plot variable.
        """
        self.__values[variable] -= value



       ###--- From this point, plot variables are created in order to work with them by using the models ---###
             ###--- They are created on the PLOT_VARS list, imported from variables file ---###

###############################################################################################################
####################################################### IDs #############################################################
###############################################################################################################

    @property
    def inventory_id(self):
        return self.__values['INVENTORY_ID']

    @property
    def plot_id(self):
        return self.__values['PLOT_ID']  

###############################################################################################################
##############################################  Plot general information  ###############################################
###############################################  variables to AREA_VARS  ################################################
###############################################################################################################

    def from_plot_to_area(plot):
        """
        Function that send information about study area from PLOT_VARS to AREA_VARS list.
        Due to SIMANFOR inventory structure, it's neccesary to import that information with plot variables; if not, 
        the process will be very difficult and maybe it will require to add a new sheet on hte initial inventory.
        Using that function, we can change the location of the information and then delete the information from 
        PLOT_VARS list with eliminate_from_plot function, in order to show it at the "Description" sheet 
        instead of "Plots" sheet.
        """

        # Change variables information from Plot to Area, being Area variables a dictionary
        if not isinstance(Area.martonne_2020, dict):       

            Area.plot_type = {}
            Area.plot_area = {}
            Area.province = {}
            Area.study_area = {}
            Area.municipality = {}
            Area.forest = {}
            Area.prov_region = {}
            Area.main_specie = {}
            #if plot.specie_ifn_id != '' and plot.specie_ifn_id != 0:
            Area.specie_ifn_id = {}
            Area.slope = {}
            Area.aspect = {}
            Area.continentality = {}
            Area.altitude = {}
            Area.longitude = {}
            Area.latitude = {}   
            Area.aa_rainfall = {}
            Area.ma_temperature = {}
            Area.september_rain = {}
            Area.september_temp = {}
            Area.november_rain = {}
            Area.november_temp = {}
            Area.martonne = {}
            Area.martonne_2020 = {}
            Area.martonne_2040 = {}
            Area.martonne_2060 = {}                
            Area.martonne_2080 = {}   
            # Cistus ladanifer special variables    
            Area.tr = {}  
            Area.rain_as = {}  
            Area.tmin_so = {} 
            Area.tmin_on = {}  
            Area.tmin_ond = {}  
            Area.tmmin_oct = {}  
            Area.tsum_mean_so = {}  
            Area.tsum_mmin_so = {}  
            Area.tsum_mmin_on = {} 
            Area.tsum_mmin_sond = {}  


        Area.plot_type[plot.plot_id] = plot.plot_type.upper()  # always like this to reduce possible mistakes on technical ingrowth selection
        Area.plot_area[plot.plot_id] = plot.plot_area
        Area.province[plot.plot_id] = plot.province
        Area.study_area[plot.plot_id] = plot.study_area
        Area.municipality[plot.plot_id] = plot.municipality
        Area.forest[plot.plot_id] = plot.forest
        Area.prov_region[plot.plot_id] = plot.prov_region
        Area.main_specie[plot.plot_id] = plot.main_specie
        #if plot.specie_ifn_id != '' and plot.specie_ifn_id != 0:
        Area.specie_ifn_id[plot.plot_id] = plot.specie_ifn_id
        Area.slope[plot.plot_id] = plot.slope
        Area.aspect[plot.plot_id] = plot.aspect
        Area.continentality[plot.plot_id] = plot.continentality
        Area.altitude[plot.plot_id] = plot.altitude
        Area.longitude[plot.plot_id] = plot.longitude
        Area.latitude[plot.plot_id] = plot.latitude   
        Area.aa_rainfall[plot.plot_id] = plot.aa_rainfall
        Area.ma_temperature[plot.plot_id] = plot.ma_temperature
        Area.september_rain[plot.plot_id] = plot.september_rain
        Area.september_temp[plot.plot_id] = plot.september_temp
        Area.november_rain[plot.plot_id] = plot.november_rain
        Area.november_temp[plot.plot_id] = plot.november_temp
        Area.martonne[plot.plot_id] = plot.martonne
        Area.martonne_2020[plot.plot_id] = plot.martonne_2020
        Area.martonne_2040[plot.plot_id] = plot.martonne_2040
        Area.martonne_2060[plot.plot_id] = plot.martonne_2060                
        Area.martonne_2080[plot.plot_id] = plot.martonne_2080   
        # Cistus ladanifer special variables    
        Area.tr[plot.plot_id] = plot.tr
        Area.rain_as[plot.plot_id] = plot.rain_as
        Area.tmin_so[plot.plot_id] = plot.tmin_so
        Area.tmin_on[plot.plot_id] = plot.tmin_on
        Area.tmin_ond[plot.plot_id] = plot.tmin_ond
        Area.tmmin_oct[plot.plot_id] = plot.tmmin_oct
        Area.tsum_mean_so[plot.plot_id] = plot.tsum_mean_so
        Area.tsum_mmin_so[plot.plot_id] = plot.tsum_mmin_so
        Area.tsum_mmin_on[plot.plot_id] = plot.tsum_mmin_on
        Area.tsum_mmin_sond[plot.plot_id] = plot.tsum_mmin_sond


    def eliminate_from_plot():
        """
        Function to eliminate information from Plot already copied on Area.
        """

        # delete variables from Plot to not show it at the "Plots" sheet
        PLOT_VARS.remove('PLOT_TYPE')
        PLOT_VARS.remove('PLOT_AREA')
        PLOT_VARS.remove('PROVINCE')
        PLOT_VARS.remove('STUDY_AREA')
        PLOT_VARS.remove('MUNICIPALITY')
        PLOT_VARS.remove('FOREST')
        PLOT_VARS.remove('PROV_REGION')
        PLOT_VARS.remove('MAIN_SPECIE')
        PLOT_VARS.remove('SPECIE_IFN_ID')
        PLOT_VARS.remove('SLOPE')
        PLOT_VARS.remove('ASPECT')
        PLOT_VARS.remove('CONTINENTALITY')
        PLOT_VARS.remove('ALTITUDE')
        PLOT_VARS.remove('LONGITUDE')
        PLOT_VARS.remove('LATITUDE')
        PLOT_VARS.remove('AA_RAINFALL')
        PLOT_VARS.remove('MA_TEMPERATURE')
        PLOT_VARS.remove('SEPTEMBER_RAIN')
        PLOT_VARS.remove('SEPTEMBER_TEMP')
        PLOT_VARS.remove('NOVEMBER_RAIN')
        PLOT_VARS.remove('NOVEMBER_TEMP')
        PLOT_VARS.remove('MARTONNE')
        PLOT_VARS.remove('MARTONNE_2020')
        PLOT_VARS.remove('MARTONNE_2040')
        PLOT_VARS.remove('MARTONNE_2060')
        PLOT_VARS.remove('MARTONNE_2080')
        # Cistus ladanifer special variables    
        PLOT_VARS.remove('TR')  # Treatment_type  # dummy variable (equal to 1 if the scrubland was developed on burned area and equal to 0 if the scrubland was developed on cleared area)
        PLOT_VARS.remove('RAIN_AS')  # Rainfall_AS  # sum of the total precipitation of August and September (mm)
        PLOT_VARS.remove('TMIN_SO')  # Tmin_SO  # is the sum of the minimum absolute temperature of September and October (ºC)
        PLOT_VARS.remove('TMIN_ON')  # Tmin_ON  # is the sum of the minimum absolute temperature of October and November (ºC)
        PLOT_VARS.remove('TMIN_OND')  # Tmin_OND  # is the sum of the minimum absolute temperature of October, November and December (ºC)
        PLOT_VARS.remove('TMMIN_OCT')  # Tmmin_O  # mean minimum temperature of October (ºC)
        PLOT_VARS.remove('TSUM_MEAN_SO')  # Tsum_mean_SO  # is the sum of the mean temperatures of September and October (ºC)
        PLOT_VARS.remove('TSUM_MMIN_SO')  # Tsum_mmin_SO # is the sum of the mean minimum temperatures of September and October (ºC)
        PLOT_VARS.remove('TSUM_MMIN_ON')  # Tsum_mmin_ON  # is the sum of the mean minimum temperatures of October and November (ºC)
        PLOT_VARS.remove('TSUM_MMIN_SOND')  # Tsum_mmin_SOND  # is the sum of the mean minimum temperatures of September, October, November and December (ºC)


    @property
    def plot_type(self):
        return self.__values['PLOT_TYPE']

    @property
    def plot_area(self):
        return self.__values['PLOT_AREA']

    @property
    def province(self):
        return self.__values['PROVINCE']

    @property
    def study_area(self):
        return self.__values['STUDY_AREA']

    @property
    def municipality(self):
        return self.__values['MUNICIPALITY']

    @property
    def forest(self):
        return self.__values['FOREST']

    @property
    def prov_region(self):
        return self.__values['PROV_REGION']

    @property
    def main_specie(self):
        return self.__values['MAIN_SPECIE']

    @property
    def specie_ifn_id(self):
        return self.__values['SPECIE_IFN_ID']

    @property
    def id_sp1(self):
        return self.__values['ID_SP1']

    @property
    def id_sp2(self):
        return self.__values['ID_SP2']        

    @property
    def slope(self):
        return self.__values['SLOPE']

    @property
    def aspect(self):
        return self.__values['ASPECT']

    @property
    def continentality(self):
        return self.__values['CONTINENTALITY']
 
    @property
    def altitude(self):
        return self.__values['ALTITUDE']

    @property
    def longitude(self):
        return self.__values['LONGITUDE']

    @property
    def latitude(self):
        return self.__values['LATITUDE']

    @property
    def aa_rainfall(self):
        return self.__values['AA_RAINFALL']

    @property
    def ma_temperature(self):
        return self.__values['MA_TEMPERATURE']        

    @property
    def september_rain(self):
        return self.__values['SEPTEMBER_RAIN']

    @property
    def september_temp(self):
        return self.__values['SEPTEMBER_TEMP']

    @property
    def november_rain(self):
        return self.__values['NOVEMBER_RAIN']

    @property
    def november_temp(self):
        return self.__values['NOVEMBER_TEMP']

    @property
    def martonne(self):
        return self.__values['MARTONNE']
   
    @property
    def martonne_2020(self):
        return self.__values['MARTONNE_2020']

    @property
    def martonne_2040(self):
        return self.__values['MARTONNE_2040']

    @property
    def martonne_2060(self):
        return self.__values['MARTONNE_2060']

    @property
    def martonne_2080(self):
        return self.__values['MARTONNE_2080']

    @property
    def year(self):
        return self.__values['YEAR']

    @property
    def tr(self):
        return self.__values['TR']

    @property
    def rain_as(self):
        return self.__values['RAIN_AS']

    @property
    def tmin_so(self):
        return self.__values['TMIN_SO']

    @property
    def tmin_on(self):
        return self.__values['TMIN_ON']  

    @property
    def tmin_ond(self):
        return self.__values['TMIN_OND']
   
    @property
    def tmmin_oct(self):
        return self.__values['TMMIN_OCT']

    @property
    def tsum_mean_so(self):
        return self.__values['TSUM_MEAN_SO']

    @property
    def tsum_mmin_so(self):
        return self.__values['TSUM_MMIN_SO']

    @property
    def tsum_mmin_on(self):
        return self.__values['TSUM_MMIN_ON']     

    @property
    def tsum_mmin_sond(self):
        return self.__values['TSUM_MMIN_SOND'] 

###############################################################################################################
############################################  Basic plot variables measured  ############################################
###############################################################################################################

    @property
    def expan(self):
        return self.__values['EXPAN']

    @property
    def age(self):
        return self.__values['AGE']

    @property
    def sp1_proportion(self):
        return self.__values['SP1_PROPORTION']

    @property
    def sp2_proportion(self):
        return self.__values['SP2_PROPORTION']

    @property
    def density(self):
        return self.__values['DENSITY']

    @property
    def density_sp1(self):
        return self.__values['DENSITY_SP1']

    @property
    def density_sp2(self):
        return self.__values['DENSITY_SP2']

    @property
    def density_cut_volume(self):
        return self.__values['DENSITY_CUT_VOLUME']

    @property
    def dead_density(self):
        return self.__values['DEAD_DENSITY']

    @property
    def ing_density(self):
        return self.__values['ING_DENSITY']

###############################################################################################################
#####################################  Basic plot variables calculated - basal area  ####################################
###############################################################################################################

    @property
    def basal_area(self):
        return self.__values['BASAL_AREA']

    @property
    def basal_area_sp1(self):
        return self.__values['BASAL_AREA_SP1']

    @property
    def basal_area_sp2(self):
        return self.__values['BASAL_AREA_SP2']

    @property
    def ba_max(self):
        return self.__values['BA_MAX']

    @property
    def ba_max_sp1(self):
        return self.__values['BA_MAX_SP1']

    @property
    def ba_max_sp2(self):
        return self.__values['BA_MAX_SP2']        

    @property
    def ba_min(self):
        return self.__values['BA_MIN']

    @property
    def ba_min_sp1(self):
        return self.__values['BA_MIN_SP1']

    @property
    def ba_min_sp2(self):
        return self.__values['BA_MIN_SP2']

    @property
    def mean_ba(self):
        return self.__values['MEAN_BA']

    @property
    def mean_ba_sp1(self):
        return self.__values['MEAN_BA_SP1']

    @property
    def mean_ba_sp2(self):
        return self.__values['MEAN_BA_SP2']

    @property
    def ba_cut_volume(self):
        return self.__values['BA_CUT_VOLUME']

    @property
    def dead_ba(self):
        return self.__values['DEAD_BA']

    @property
    def ing_ba(self):
        return self.__values['ING_BA']

###############################################################################################################
######################################  Basic plot variables calculated - diameter  #####################################
###############################################################################################################

    @property
    def dbh_max(self):
        return self.__values['DBH_MAX']

    @property
    def dbh_max_sp1(self):
        return self.__values['DBH_MAX_SP1']

    @property
    def dbh_max_sp2(self):
        return self.__values['DBH_MAX_SP2']

    @property
    def dbh_min(self):
        return self.__values['DBH_MIN']

    @property
    def dbh_min_sp1(self):
        return self.__values['DBH_MIN_SP1']

    @property
    def dbh_min_sp2(self):
        return self.__values['DBH_MIN_SP2']

    @property
    def mean_dbh(self):
        return self.__values['MEAN_DBH']

    @property
    def mean_dbh_sp1(self):
        return self.__values['MEAN_DBH_SP1']

    @property
    def mean_dbh_sp2(self):
        return self.__values['MEAN_DBH_SP2']

    @property
    def qm_dbh(self):
        return self.__values['QM_DBH']

    @property
    def qm_dbh_sp1(self):
        return self.__values['QM_DBH_SP1']

    @property
    def qm_dbh_sp2(self):
        return self.__values['QM_DBH_SP2']

    @property
    def dominant_dbh(self):
        return self.__values['DOMINANT_DBH']

    @property
    def dominant_dbh_sp1(self):
        return self.__values['DOMINANT_DBH_SP1']

    @property
    def dominant_dbh_sp2(self):
        return self.__values['DOMINANT_DBH_SP2']                

    @property
    def dominant_section(self):
        return self.__values['DOMINANT_SECTION']

    @property
    def dominant_section_sp1(self):
        return self.__values['DOMINANT_SECTION_SP1']

    @property
    def dominant_section_sp2(self):
        return self.__values['DOMINANT_SECTION_SP2']        

###############################################################################################################
######################################## Basic plot variables calculated - height #######################################
###############################################################################################################

    @property
    def h_max(self):
        return self.__values['H_MAX']

    @property
    def h_max_sp1(self):
        return self.__values['H_MAX_SP1']

    @property
    def h_max_sp2(self):
        return self.__values['H_MAX_SP2']

    @property
    def h_min(self):
        return self.__values['H_MIN']

    @property
    def h_min_sp1(self):
        return self.__values['H_MIN_SP1']

    @property
    def h_min_sp2(self):
        return self.__values['H_MIN_SP2']

    @property
    def mean_h(self):
        return self.__values['MEAN_H']

    @property
    def mean_h_sp1(self):
        return self.__values['MEAN_H_SP1']

    @property
    def mean_h_sp2(self):
        return self.__values['MEAN_H_SP2']

    @property
    def dominant_h(self):
        return self.__values['DOMINANT_H']

    @property
    def dominant_h_sp1(self):
        return self.__values['DOMINANT_H_SP1']

    @property
    def dominant_h_sp2(self):
        return self.__values['DOMINANT_H_SP2']

###############################################################################################################
#######################################  Basic plot variables calculated - crown  #######################################
###############################################################################################################

    @property
    def crown_mean_d(self):
        return self.__values['CROWN_MEAN_D']

    @property
    def crown_mean_d_sp1(self):
        return self.__values['CROWN_MEAN_D_SP1']

    @property
    def crown_mean_d_sp2(self):
        return self.__values['CROWN_MEAN_D_SP2']

    @property
    def crown_dom_d(self):
        return self.__values['CROWN_DOM_D']

    @property
    def crown_dom_d_sp1(self):
        return self.__values['CROWN_DOM_D_SP1']

    @property
    def crown_dom_d_sp2(self):
        return self.__values['CROWN_DOM_D_SP2']

    @property
    def canopy_cover(self):
        return self.__values['CANOPY_COVER']

    @property
    def canopy_cover_sp1(self):
        return self.__values['CANOPY_COVER_SP1']

    @property
    def canopy_cover_sp2(self):
        return self.__values['CANOPY_COVER_SP2']        

    @property
    def canopy_vol(self):
        return self.__values['CANOPY_VOL']

    @property
    def canopy_vol_sp1(self):
        return self.__values['CANOPY_VOL_SP1']

    @property
    def canopy_vol_sp2(self):
        return self.__values['CANOPY_VOL_SP2']    

###############################################################################################################
######################################## Basic plot variables calculated - plot  ########################################
###############################################################################################################

    @property
    def slenderness_mean(self):
        return self.__values['SLENDERNESS_MEAN']

    @property
    def slenderness_dom(self):
        return self.__values['SLENDERNESS_DOM']

    @property
    def reineke(self):
        return self.__values['REINEKE']

    @property
    def reineke_sp1(self):
        return self.__values['REINEKE_SP1']

    @property
    def reineke_sp2(self):
        return self.__values['REINEKE_SP2']

    @property
    def reineke_max(self):
        return self.__values['REINEKE_MAX']

    @property
    def reineke_max_sp1(self):
        return self.__values['REINEKE_MAX_SP1']

    @property
    def reineke_max_sp2(self):
        return self.__values['REINEKE_MAX_SP2']

    @property
    def reineke_value(self):
        return self.__values['REINEKE_VALUE']

    @property
    def hart(self):
        return self.__values['HART']

    @property
    def hart_staggered(self):
        return self.__values['HART_STAGGERED']

    @property
    def si(self):
        return self.__values['SI']

    @property
    def ref_si_age(self):
        return self.__values['REF_SI_AGE']

    @property
    def hegyi_radius(self):
        return self.__values['HEGYI_RADIUS']

###############################################################################################################
########################################### Plot variables calculated - volume ##########################################
###############################################################################################################

    @property
    def vol(self):
        return self.__values['VOL']

    @property
    def bole_vol(self):
        return self.__values['BOLE_VOL']

    @property
    def bark_vol(self):
        return self.__values['BARK_VOL']

    @property
    def vol_sp1(self):
        return self.__values['VOL_SP1']

    @property
    def bole_vol_sp1(self):
        return self.__values['BOLE_VOL_SP1']

    @property
    def bark_vol_sp1(self):
        return self.__values['BARK_VOL_SP1']

    @property
    def vol_sp2(self):
        return self.__values['VOL_SP2']

    @property
    def bole_vol_sp2(self):
        return self.__values['BOLE_VOL_SP2']

    @property
    def bark_vol_sp2(self):
        return self.__values['BARK_VOL_SP2']

    @property
    def vol_cut_volume(self):
        return self.__values['VOL_CUT_VOLUME']

    @property
    def dead_vol(self):
        return self.__values['DEAD_VOL']

    @property
    def ing_vol(self):
        return self.__values['ING_VOL']

###############################################################################################################
########################################## Plot variables calculated - biomass ##########################################
###############################################################################################################

    @property
    def wsw(self):
        return self.__values['WSW']

    @property
    def wsb(self):
        return self.__values['WSB']

    @property
    def wsb(self):
        return self.__values['WSWB']

    @property
    def wthickb(self):
        return self.__values['WTHICKB']

    @property
    def wstb(self):
        return self.__values['WSTB']

    @property
    def wb2_7(self):
        return self.__values['WB2_7']

    @property
    def wb2_t(self):
        return self.__values['WB2_T']

    @property
    def wthinb(self):
        return self.__values['WTHINB']

    @property
    def wb05(self):
        return self.__values['WB05']

    @property
    def wb05_7(self):
        return self.__values['WB05_7']

    @property
    def wb0_2(self):
        return self.__values['WB0_2']

    @property
    def wdb(self):
        return self.__values['WDB']

    @property
    def wl(self):
        return self.__values['WL']

    @property
    def wtbl(self):
        return self.__values['WTBL']

    @property
    def wbl0_7(self):
        return self.__values['WBL0_7']

    @property
    def wr(self):
        return self.__values['WR']

    @property
    def wt(self):
        return self.__values['WT']

    @property
    def wt_cut_volume(self):
        return self.__values['WT_CUT_VOLUME']

    @property
    def dead_wt(self):
        return self.__values['DEAD_WT']

    @property
    def ing_wt(self):
        return self.__values['ING_WT']

###############################################################################################################
########################################## Plot variables calculated - biomass ##########################################
###############################################################################################################

    @property
    def wsw_sp1(self):
        return self.__values['WSW_SP1']

    @property
    def wsb_sp1(self):
        return self.__values['WSB_SP1']

    @property
    def wsb_sp1(self):
        return self.__values['WSWB_SP1']

    @property
    def wthickb_sp1(self):
        return self.__values['WTHICKB_SP1']

    @property
    def wstb_sp1(self):
        return self.__values['WSTB_SP1']

    @property
    def wb2_7_sp1(self):
        return self.__values['WB2_7_SP1']

    @property
    def wb2_t_sp1(self):
        return self.__values['WB2_T_SP1']

    @property
    def wthinb_sp1(self):
        return self.__values['WTHINB_SP1']

    @property
    def wb05_sp1(self):
        return self.__values['WB05_SP1']

    @property
    def wb05_7_sp1(self):
        return self.__values['WB05_7_SP1']

    @property
    def wb0_2_sp1(self):
        return self.__values['WB0_2_SP1']

    @property
    def wdb_sp1(self):
        return self.__values['WDB_SP1']

    @property
    def wl_sp1(self):
        return self.__values['WL_SP1']

    @property
    def wtbl_sp1(self):
        return self.__values['WTBL_SP1']

    @property
    def wbl0_7_sp1(self):
        return self.__values['WBL0_7_SP1']

    @property
    def wr_sp1(self):
        return self.__values['WR_SP1']

    @property
    def wt_sp1(self):
        return self.__values['WT_SP1']


    @property
    def wsw_sp2(self):
        return self.__values['WSW_SP2']

    @property
    def wsb_sp2(self):
        return self.__values['WSB_SP2']

    @property
    def wsb_sp2(self):
        return self.__values['WSWB_SP2']

    @property
    def wthickb_sp2(self):
        return self.__values['WTHICKB_SP2']

    @property
    def wstb_sp2(self):
        return self.__values['WSTB_SP2']

    @property
    def wb2_7_sp2(self):
        return self.__values['WB2_7_SP2']

    @property
    def wb2_t_sp2(self):
        return self.__values['WB2_T_SP2']

    @property
    def wthinb_sp2(self):
        return self.__values['WTHINB_SP2']

    @property
    def wb05_sp2(self):
        return self.__values['WB05_SP2']

    @property
    def wb05_7_sp2(self):
        return self.__values['WB05_7_SP2']

    @property
    def wb0_2_sp2(self):
        return self.__values['WB0_2_SP2']

    @property
    def wdb_sp2(self):
        return self.__values['WDB_SP2']

    @property
    def wl_sp2(self):
        return self.__values['WL_SP2']

    @property
    def wtbl_sp2(self):
        return self.__values['WTBL_SP2']

    @property
    def wbl0_7_sp2(self):
        return self.__values['WBL0_7_SP2']

    @property
    def wr_sp2(self):
        return self.__values['WR_SP2']

    @property
    def wt_sp2(self):
        return self.__values['WT_SP2']

###############################################################################################################
######################################### Plot variables calculated - wood uses #########################################
###############################################################################################################
    
    @property
    def unwinding(self):
        return self.__values['UNWINDING']

    @property
    def veneer(self):
        return self.__values['VENEER']

    @property
    def saw_big(self):
        return self.__values['SAW_BIG']

    @property
    def saw_small(self):
        return self.__values['SAW_SMALL']

    @property
    def saw_canter(self):
        return self.__values['SAW_CANTER']

    @property
    def post(self):
        return self.__values['POST']

    @property
    def stake(self):
        return self.__values['STAKE']

    @property
    def chips(self):
        return self.__values['CHIPS'] 

    @property
    def unwinding_sp1(self):
        return self.__values['UNWINDING_SP1']

    @property
    def veneer_sp1(self):
        return self.__values['VENEER_SP1']

    @property
    def saw_big_sp1(self):
        return self.__values['SAW_BIG_SP1']

    @property
    def saw_small_sp1(self):
        return self.__values['SAW_SMALL_SP1']

    @property
    def saw_canter_sp1(self):
        return self.__values['SAW_CANTER_SP1']

    @property
    def post_sp1(self):
        return self.__values['POST_SP1']

    @property
    def stake_sp1(self):
        return self.__values['STAKE_SP1']

    @property
    def chips_sp1(self):
        return self.__values['CHIPS_SP1'] 

    @property
    def unwinding_sp2(self):
        return self.__values['UNWINDING_SP2']

    @property
    def veneer_sp2(self):
        return self.__values['VENEER_SP2']

    @property
    def saw_big_sp2(self):
        return self.__values['SAW_BIG_SP2']

    @property
    def saw_small_sp2(self):
        return self.__values['SAW_SMALL_SP2']

    @property
    def saw_canter_sp2(self):
        return self.__values['SAW_CANTER_SP2']

    @property
    def post_sp2(self):
        return self.__values['POST_SP2']

    @property
    def stake_sp2(self):
        return self.__values['STAKE_SP2']

    @property
    def chips_sp2(self):
        return self.__values['CHIPS_SP2']

###############################################################################################################
####################################  Species diversity indexes ###############################################
###############################################################################################################

        @property
        def shannon(self):
            return self.__values['SHANNON']

        @property
        def simpson(self):
            return self.__values['SIMPSON']

        @property
        def margalef(self):
            return self.__values['MARGALEF']

        @property
        def pielou(self):
            return self.__values['PIELOU']

###############################################################################################################
#######################################  Pinus pinea special variables ##################################################
###############################################################################################################

    @property
    def w_cork(self):
        return self.__values['W_CORK']

    @property
    def total_w_debark(self):
        return self.__values['TOTAL_W_DEBARK']

    @property
    def total_v_debark(self):
        return self.__values['TOTAL_V_DEBARK']

###############################################################################################################
#######################################  Pinus pinea special variables ##################################################
###############################################################################################################

    @property
    def all_cones(self):
        return self.__values['ALL_CONES']

    @property
    def sound_cones(self):
        return self.__values['SOUND_CONES']

    @property
    def sound_seeds(self):
        return self.__values['SOUND_SEEDS']

    @property
    def w_sound_cones(self):
        return self.__values['W_SOUND_CONES']

    @property
    def w_all_cones(self):
        return self.__values['W_ALL_CONES']

###############################################################################################################
########################################  Mushrooms special variables ###################################################
###############################################################################################################

    @property
    def edible_mush(self):
        return self.__values['EDIBLE_MUSH']

    @property
    def marketed_mush(self):
        return self.__values['MARKETED_MUSH']

    @property
    def marketed_lactarius(self):
        return self.__values['MARKETED_LACTARIUS']

    @property
    def mushroom_productivity(self):
        return self.__values['MUSHROOM_PRODUCTIVITY']

###############################################################################################################
########################################  Mushrooms special variables ###################################################
###############################################################################################################

    @property
    def time_at(self):
        return self.__values['TIME_AT']

    @property
    def b_edulis(self):
        return self.__values['B_EDULIS']

    @property
    def mycod(self):
        return self.__values['MYCOD']

    @property
    def mycop(self):
        return self.__values['MYCOP']

    @property
    def saprod(self):
        return self.__values['SAPROD']

    @property
    def saprop(self):
        return self.__values['SAPROP']

###############################################################################################################
############################## Auxiliar variables for future models - 11/08/2023 ##############################
###############################################################################################################

    @property
    def plot_var1(self):
        return self.__values['PLOT_VAR1']

    @property
    def plot_var2(self):
        return self.__values['PLOT_VAR2']

    @property
    def plot_var3(self):
        return self.__values['PLOT_VAR3']

    @property
    def plot_var4(self):
        return self.__values['PLOT_VAR4']

    @property
    def plot_var5(self):
        return self.__values['PLOT_VAR5']

    @property
    def plot_var6(self):
        return self.__values['PLOT_VAR6']

    @property
    def plot_var7(self):
        return self.__values['PLOT_VAR7']

    @property
    def plot_var8(self):
        return self.__values['PLOT_VAR8']

    @property
    def plot_var9(self):
        return self.__values['PLOT_VAR9']

    @property
    def plot_var10(self):
        return self.__values['PLOT_VAR10']

    @property
    def plot_var11(self):
        return self.__values['PLOT_VAR11']

    @property
    def plot_var12(self):
        return self.__values['PLOT_VAR12']

    @property
    def plot_var13(self):
        return self.__values['PLOT_VAR13']

    @property
    def plot_var14(self):
        return self.__values['PLOT_VAR14']

    @property
    def plot_var15(self):
        return self.__values['PLOT_VAR15']

    @property
    def plot_var16(self):
        return self.__values['PLOT_VAR16']

    @property
    def plot_var17(self):
        return self.__values['PLOT_VAR17']

    @property
    def plot_var18(self):
        return self.__values['PLOT_VAR18']

    @property
    def plot_var19(self):
        return self.__values['PLOT_VAR19']

    @property
    def plot_var20(self):
        return self.__values['PLOT_VAR20']

    @property
    def plot_var21(self):
        return self.__values['PLOT_VAR21']

    @property
    def plot_var22(self):
        return self.__values['PLOT_VAR22']

    @property
    def plot_var23(self):
        return self.__values['PLOT_VAR23']

    @property
    def plot_var24(self):
        return self.__values['PLOT_VAR24']

    @property
    def plot_var25(self):
        return self.__values['PLOT_VAR25']

    @property
    def plot_var26(self):
        return self.__values['PLOT_VAR26']

    @property
    def plot_var27(self):
        return self.__values['PLOT_VAR27']

    @property
    def plot_var28(self):
        return self.__values['PLOT_VAR28']

    @property
    def plot_var29(self):
        return self.__values['PLOT_VAR29']

    @property
    def plot_var30(self):
        return self.__values['PLOT_VAR30']

    @property
    def plot_var31(self):
        return self.__values['PLOT_VAR31']

    @property
    def plot_var32(self):
        return self.__values['PLOT_VAR32']

    @property
    def plot_var33(self):
        return self.__values['PLOT_VAR33']

    @property
    def plot_var34(self):
        return self.__values['PLOT_VAR34']

    @property
    def plot_var35(self):
        return self.__values['PLOT_VAR35']

    @property
    def plot_var36(self):
        return self.__values['PLOT_VAR36']

    @property
    def plot_var37(self):
        return self.__values['PLOT_VAR37']

    @property
    def plot_var38(self):
        return self.__values['PLOT_VAR38']

    @property
    def plot_var39(self):
        return self.__values['PLOT_VAR39']

    @property
    def plot_var40(self):
        return self.__values['PLOT_VAR40']

    @property
    def plot_var41(self):
        return self.__values['PLOT_VAR41']

    @property
    def plot_var42(self):
        return self.__values['PLOT_VAR42']

    @property
    def plot_var43(self):
        return self.__values['PLOT_VAR43']

    @property
    def plot_var44(self):
        return self.__values['PLOT_VAR44']

    @property
    def plot_var45(self):
        return self.__values['PLOT_VAR45']

    @property
    def plot_var46(self):
        return self.__values['PLOT_VAR46']

    @property
    def plot_var47(self):
        return self.__values['PLOT_VAR47']

    @property
    def plot_var48(self):
        return self.__values['PLOT_VAR48']

    @property
    def unwinding_sp3(self):
        return self.__values['UNWINDING_SP3']

    @property
    def veneer_sp3(self):
        return self.__values['VENEER_SP3']

    @property
    def saw_big_sp3(self):
        return self.__values['SAW_BIG_SP3']

    @property
    def saw_small_sp3(self):
        return self.__values['SAW_SMALL_SP3']

    @property
    def saw_canter_sp3(self):
        return self.__values['SAW_CANTER_SP3']

    @property
    def post_sp3(self):
        return self.__values['POST_SP3']

    @property
    def stake_sp3(self):
        return self.__values['STAKE_SP3']

    @property
    def chips_sp3(self):
        return self.__values['CHIPS_SP3']

    @property
    def vol_sp3(self):
        return self.__values['VOL_SP3']

    @property
    def bole_vol_sp3(self):
        return self.__values['BOLE_VOL_SP3']

    @property
    def bark_vol_sp3(self):
        return self.__values['BARK_VOL_SP3']

    @property
    def carbon_stem_sp1(self):
        return self.__values['CARBON_STEM_SP1']

    @property
    def carbon_branches_sp1(self):
        return self.__values['CARBON_BRANCHES_SP1']

    @property
    def carbon_roots_sp1(self):
        return self.__values['CARBON_ROOTS_SP1']

    @property
    def carbon_sp1(self):
        return self.__values['CARBON_SP1']

    @property
    def carbon_stem_sp2(self):
        return self.__values['CARBON_STEM_SP2']

    @property
    def carbon_branches_sp2(self):
        return self.__values['CARBON_BRANCHES_SP2']

    @property
    def carbon_roots_sp2(self):
        return self.__values['CARBON_ROOTS_SP2']

    @property
    def carbon_sp2(self):
        return self.__values['CARBON_SP2']

    @property
    def carbon_stem_sp3(self):
        return self.__values['CARBON_STEM_SP3']

    @property
    def carbon_branches_sp3(self):
        return self.__values['CARBON_BRANCHES_SP3']

    @property
    def carbon_roots_sp3(self):
        return self.__values['CARBON_ROOTS_SP3']

    @property
    def carbon_sp3(self):
        return self.__values['CARBON_SP3']

    @property
    def ws_sp3(self):
        return self.__values['WS_SP3']

    @property
    def wb_sp3(self):
        return self.__values['WB_SP3']

    @property
    def wr_sp3(self):
        return self.__values['WR_SP3']

    @property
    def wt_sp3(self):
        return self.__values['WT_SP3']

    @property
    def ws_sp2(self):
        return self.__values['WS_SP2']

    @property
    def wb_sp2(self):
        return self.__values['WB_SP2']

    @property
    def ws_sp1(self):
        return self.__values['WS_SP1']

    @property
    def wb_sp1(self):
        return self.__values['WB_SP1']

    @property
    def carbon_stem(self):
        return self.__values['CARBON_STEM']

    @property
    def carbon_branches(self):
        return self.__values['CARBON_BRANCHES']

    @property
    def carbon_roots(self):
        return self.__values['CARBON_ROOTS']

    @property
    def wb(self):
        return self.__values['WB']

    @property
    def ws(self):
        return self.__values['WS']

    @property
    def zpcum9(self):
        return self.__values['ZPCUM9']

    @property
    def zpcum8(self):
        return self.__values['ZPCUM8']

    @property
    def zpcum7(self):
        return self.__values['ZPCUM7']

    @property
    def zpcum6(self):
        return self.__values['ZPCUM6']

    @property
    def zpcum5(self):
        return self.__values['ZPCUM5']

    @property
    def zpcum4(self):
        return self.__values['ZPCUM4']

    @property
    def zpcum3(self):
        return self.__values['ZPCUM3']

    @property
    def zpcum2(self):
        return self.__values['ZPCUM2']

    @property
    def zpcum1(self):
        return self.__values['ZPCUM1']

    @property
    def zq95(self):
        return self.__values['ZQ95']

    @property
    def zq90(self):
        return self.__values['ZQ90']

    @property
    def zq85(self):
        return self.__values['ZQ85']

    @property
    def zq80(self):
        return self.__values['ZQ80']

    @property
    def zq75(self):
        return self.__values['ZQ75']

    @property
    def zq70(self):
        return self.__values['ZQ70']

    @property
    def zq65(self):
        return self.__values['ZQ65']

    @property
    def zq60(self):
        return self.__values['ZQ60']

    @property
    def zq55(self):
        return self.__values['ZQ55']

    @property
    def zq50(self):
        return self.__values['ZQ50']

    @property
    def zq45(self):
        return self.__values['ZQ45']

    @property
    def zq40(self):
        return self.__values['ZQ40']

    @property
    def zq35(self):
        return self.__values['ZQ35']

    @property
    def zq30(self):
        return self.__values['ZQ30']

    @property
    def zq25(self):
        return self.__values['ZQ25']

    @property
    def zq20(self):
        return self.__values['ZQ20']

    @property
    def zq15(self):
        return self.__values['ZQ15']

    @property
    def zq10(self):
        return self.__values['ZQ10']

    @property
    def zq5(self):
        return self.__values['ZQ5']

    @property
    def pzabove2(self):
        return self.__values['PZABOVE2']

    @property
    def pzabovemean(self):
        return self.__values['PZABOVEZMEAN']

    @property
    def zentropy(self):
        return self.__values['ZENTROPY']

    @property
    def zkurt(self):
        return self.__values['ZKURT']

    @property
    def zskew(self):
        return self.__values['ZSKEW']

    @property
    def zsd(self):
        return self.__values['ZSD']

    @property
    def zmean(self):
        return self.__values['ZMEAN']

    @property
    def zmax(self):
        return self.__values['ZMAX']

    @property
    def sp3_n_proportion(self):
        return self.__values['SP3_N_PROPORTION']

    @property
    def sp2_n_proportion(self):
        return self.__values['SP2_N_PROPORTION']

    @property
    def sp1_n_proportion(self):
        return self.__values['SP1_N_PROPORTION']

    @property
    def basal_area_sp3(self):
        return self.__values['BASAL_AREA_SP3']

    @property
    def qm_dbh_sp3(self):
        return self.__values['QM_DBH_SP3']

    @property
    def density_sp3(self):
        return self.__values['DENSITY_SP3']

    @property
    def mean_h_sp3(self):
        return self.__values['MEAN_H_SP3']

    @property
    def h_min_sp3(self):
        return self.__values['H_MIN_SP3']

    @property
    def h_max_sp3(self):
        return self.__values['H_MAX_SP3']

    @property
    def mean_dbh_sp3(self):
        return self.__values['MEAN_DBH_SP3']

    @property
    def dbh_min_sp3(self):
        return self.__values['DBH_MIN_SP3']

    @property
    def dbh_max_sp3(self):
        return self.__values['DBH_MAX_SP3']

    @property
    def mean_ba_sp3(self):
        return self.__values['MEAN_BA_SP3']

    @property
    def ba_min_sp3(self):
        return self.__values['BA_MIN_SP3']

    @property
    def ba_max_sp3(self):
        return self.__values['BA_MAX_SP3']

    @property
    def dominant_section_sp3(self):
        return self.__values['DOMINANT_SECTION_SP3']

    @property
    def dominant_dbh_sp3(self):
        return self.__values['DOMINANT_DBH_SP3']

    @property
    def dominant_h_sp3(self):
        return self.__values['DOMINANT_H_SP3']

    @property
    def carbon_heartwood(self):
        return self.__values['CARBON_HEARTWOOD']

    @property
    def carbon_sapwood(self):
        return self.__values['CARBON_SAPWOOD']

    @property
    def carbon_bark(self):
        return self.__values['CARBON_BARK']

    @property
    def deadwood_index_cesefor_g(self):
        return self.__values['DEADWOOD_INDEX_CESEFOR_G']

    @property
    def deadwood_index_cesefor_v(self):
        return self.__values['DEADWOOD_INDEX_CESEFOR_V']

    @property
    def saw_big_liferebollo(self):
        return self.__values['SAW_BIG_LIFEREBOLLO']

    @property
    def saw_small_liferebollo(self):
        return self.__values['SAW_SMALL_LIFEREBOLLO']

    @property
    def staves_intona(self):
        return self.__values['STAVES_INTONA']

    @property
    def bottom_staves_intona(self):
        return self.__values['BOTTOM_STAVES_INTONA']

    @property
    def wood_panels_gamiz(self):
        return self.__values['WOOD_PANELS_GAMIZ']

    @property
    def mix_garcia_varona(self):
        return self.__values['MIX_GARCIA_VARONA']

    @property
    def carbon(self):
        return self.__values['CARBON']

###############################################################################################################
        
    def clone(self, plot, full=False):
        """
        Function used to clone the plot information.
        It is used on basic_engine file.
        """
        
        for variable in PLOT_VARS:
            self.__values[variable] = plot.get_value(variable)

        if full:
            for tree in plot.trees:
                tmp_tree = Tree()
                tmp_tree.clone(tree)
                self.__trees[tmp_tree.id] = tmp_tree

    def clone_by_variable(self, plot, variable: str, value):

        for variable in PLOT_VARS:
            self.__values[variable] = plot.get_value(variable)
        for tree in plot.trees:
            if tree.get_value(variable) == value:
                tmp_tree = Tree()
                tmp_tree.clone(tree)
                self.__trees[tmp_tree.id] = tmp_tree

    def get_dominant_height(self, selection_trees: list):
        """
        Function to calculate dominant height variable by using trees information.
        It is used at recalculate function on this file, and sometimes from stand models.
        """

        acumulate: float = 0
        result: float = 0

        for tree in selection_trees:
            if tree.height != 0 and tree.height != '':  # avoid trees without height values
                if acumulate + tree.expan < 100:
                    result += tree.height * tree.expan
                    acumulate += tree.expan
                else:
                    result += (100 - acumulate) * tree.height
                    acumulate = 100
        
        if acumulate != 0:
            return result / acumulate
        else:
            return 0


    def get_dominant_diameter(self, selection_trees: list):
        """
        Function to calculate dominant diameter variable by using trees information.
        It is used at recalculate function on this file, and sometimes from stand models.
        """
        
        acumulate: float = 0
        result: float = 0

        for tree in selection_trees:
            if acumulate + tree.expan < 100:
                result += tree.dbh * tree.expan
                acumulate += tree.expan
            else:
                result += (100 - acumulate) * tree.dbh
                acumulate = 100
        
        if acumulate != 0:
            return result / acumulate
        else:
            return 0
            

    def get_dominant_section(self, selection_trees: list):
        """
        Function to calculate dominant section variable by using trees information.
        It is used at recalculate function on this file, and sometimes from stand models.
        """

        acumulate: float = 0
        result: float = 0

        for tree in selection_trees:
            if acumulate + tree.expan < 100:
                result += tree.basal_area * tree.expan
                acumulate += tree.expan
            else:
                result += (100 - acumulate) * tree.basal_area
                acumulate = 100
        
        if acumulate != 0:
            return result / acumulate
        else:
            return 0


    def recalculate(self):
        """
        Function necessary to recalculate the plot variables.
        It is used when tree and plot variables are updated after an execution. Only alive trees are used inside that function.
        """

        tree_expansion: float = 0.0

        order_criteria = OrderCriteria(ASC)
        order_criteria.add_criteria('dbh')

        expansion_trees = Tree.get_sord_and_order_tree_list(self.__trees.values(), order_criteria=order_criteria)
        selection_trees = list()

        for tree in expansion_trees:  # select 100 taller trees
            if tree_expansion < 100:
                tree_expansion += tree.expan
                selection_trees.append(tree)
            else:
                break


        plot_expan = plot_expan_2 = plot_ba = plot_dbh = plot_dbh2 = max_dbh = max_h = max_ba = plot_h = 0
        min_dbh = min_ba = min_h = 9999

        for tree in self.__trees.values():

            if tree.height == 0 and tree.bal == 0 and tree.vol == 0 and tree.dbh != 0:  # skip trees created on ingrowth function
                continue

            plot_expan += tree.expan
            plot_ba += tree.basal_area*tree.expan
            plot_dbh += tree.dbh*tree.expan
            plot_dbh2 += math.pow(tree.dbh, 2)*tree.expan
            
            if tree.height != 0 and tree.height != '':  # condition to avoid lack of tree height information
                plot_h += tree.height*tree.expan
                plot_expan_2 += tree.expan  
            
                max_h = tree.height if tree.height > max_h else max_h
                min_h = tree.height if tree.height < min_h else min_h

            max_dbh = tree.dbh if tree.dbh > max_dbh else max_dbh
            min_dbh = tree.dbh if tree.dbh < min_dbh else min_dbh

            max_ba = tree.basal_area if tree.basal_area > max_ba else max_ba
            min_ba = tree.basal_area if tree.basal_area < min_ba else min_ba

        self.__values['DENSITY'] = plot_expan

        self.__values['DOMINANT_DBH'] = self.get_dominant_diameter(selection_trees)
        if plot_expan != 0:
            self.__values['MEAN_DBH'] = plot_dbh/plot_expan
            self.__values['QM_DBH'] = math.sqrt(plot_dbh2/plot_expan)     
        self.__values['DBH_MAX'] = max_dbh
        self.__values['DBH_MIN'] = min_dbh

        self.__values['DOMINANT_H'] = self.get_dominant_height(selection_trees)
        if plot_expan_2 != 0:
            self.__values['MEAN_H'] = plot_h/plot_expan_2      
        self.__values['H_MAX'] = max_h
        self.__values['H_MIN'] = min_h

        self.__values['DOMINANT_SECTION'] = self.get_dominant_section(selection_trees)        
        self.__values['BASAL_AREA'] = plot_ba/10000
        if plot_expan != 0:
            self.__values['MEAN_BA'] = plot_ba/plot_expan
        self.__values['BA_MAX'] = max_ba
        self.__values['BA_MIN'] = min_ba

        if self.dominant_dbh != '' and self.dominant_dbh != 0:
            self.__values['SLENDERNESS_DOM'] = self.dominant_h*100/self.dominant_dbh
        if self.mean_dbh != '' and self.mean_dbh != 0:
            self.__values['SLENDERNESS_MEAN'] = self.mean_h*100/self.mean_dbh        

        if self.__values['QM_DBH'] != 0:
            if 'REINEKE_VALUE' in PLOT_VARS:
                self.__values['REINEKE'] = plot_expan*math.pow(25/self.qm_dbh, self.__values['REINEKE_VALUE'])
            else:
                self.__values['REINEKE'] = plot_expan*math.pow(25/self.qm_dbh, -1.605)
        else:
            self.__values['REINEKE'] = 0

        if self.dominant_h != 0:
            self.__values['HART'] = 10000/(self.dominant_h*math.sqrt(plot_expan))
            self.__values['HART_STAGGERED'] = (10000/self.dominant_h)*math.sqrt(2/(plot_expan*math.sqrt(3)))

        return self


    def calculate_plot_from_tree(self):
        """
        Function designed to obtain the plot values from the tree information.
        Actually it is not working, is a simple translation from the last Simanfor
        """

        for tree in self.__trees.values():

            if tree is not None:
                self.__values['BASAL_AREA'] = tree.basal_area * tree.expan / 10000;
                self.__values['DOMINANT_H'] = tree.height
                self.__values['DENSITY'] = tree.expan
                self.__values['AGE'] = tree.tree_age
                self.__values['MEAN_DBH'] = tree.var_1
                self.__values['QM_DBH'] = tree.dbh
                self.__values['DOMINANT_DBH'] = tree.var_2
                self.__values['DBH_MAX'] = tree.var_4
                self.__values['DBH_MIN'] = tree.var_5
                self.__values['MEAN_H'] = tree.var_3
                self.__values['CROWN_MEAN_D'] = tree.var_6
                self.__values['CROWN_DOM_D'] = tree.lcw

                if self.qm_dbh != 0:
                    self.__values['REINEKE'] = tree.expan * math.pow(25 / self.qm_dbh, -1.605)

                if self.dominant_h != 0 and tree.expan != 0:
                    self.__values['HART'] = 10000 / (self.dominant_h * math.sqrt(tree.expan))

                if self.__values['DOMINANT_H'] != 0:
                    self.__values['CANOPY_COVER'] = math.pi * (tree.lcw * tree.lcw / 4) * tree.expan / 10000

                self.__values['VOL'] = tree.vol * tree.expan
                self.__values['BOLE_VOL'] = tree.bole_vol * tree.expan
                self.__values['BARK_VOL'] = self.__values['VOL'] - self.__values['BOLE_VOL']
            else:
                Tools.print_log_line('Tree is None', logging.ERROR)

    def debark_plot(self, w_debark, v_debark):
        """
        Function to calculate cork plot variables after a debark process.
        It is called from debark.py after to calculate tree results
        """

        self.__values['TOTAL_W_DEBARK'] = w_debark
        self.__values['TOTAL_V_DEBARK'] = v_debark        

    def cut_vars(self):
        """
        That function sets DEAD and INGROWTH plot variables as 0 when a cut is done.
        It is activated since basic_engine file, apply_harvest_model function, once plot variables are recalculated.
        """
        
        if 'DEAD_DENSITY' in PLOT_VARS and self.__values['DEAD_DENSITY'] != '':  # set it as 0 only if it was activated before; else it means that the model hasn't survive equation
            self.__values['DEAD_DENSITY'] = self.__values['DEAD_WT'] = self.__values['DEAD_BA'] = self.__values['DEAD_VOL'] = 0
        if 'ING_DENSITY' in PLOT_VARS and self.__values['ING_DENSITY'] != '':  # set it as 0 only if it was activated before; else it means that the model hasn't ingrowth equation
            self.__values['ING_DENSITY'] = self.__values['ING_WT'] = self.__values['ING_BA'] = self.__values['ING_VOL'] = 0

    def get_first(self):
        for tree in self.__trees:
            return tree
        return None

    def get_first_tree(self, variable: str, value: float):

        for tree in self.__trees:
            if tree.get_value(variable) == value:
                return tree
        return None

    def update_trees(self, variables: dict, action: int = 1):
        for tree in self.__trees.values():
            for key, value in variables.items():
                if action == 1:
                    tree.sum_value(key, value)
                elif action == 2:
                    tree.sub_value(key, value)
                else:
                    tree.set_value(key, value)

    def plot_to_json(self):

        content = dict()

        for i in range(len(PLOT_VARS)):
            content[PLOT_VARS] = self.__values[PLOT_VARS[i]]

        return content

    def print_value(self, variable, dec_pts: int = 2):
        """
        Function neccesary to print the tree values on the output.
        """

        if isinstance(self.__values[variable], float):
            return round(self.__values[variable], dec_pts)
        return self.__values[variable]

    def trees_to_json(self):

        content = dict()

        for tree in self.__trees.values():
            content[tree.id] = tree.to_json()

        return content


    def antes(self, plot, ws_summary, summary_row, dec_pts, time, first_line = None):
        """
        Function to print the plot information before a cut at the summary sheet of the output.
        """

        #
        # Información general - General information
        #
        if 'AGE' in PLOT_VARS:  # we print at the first summary column the stand age...
            ws_summary.cell(row=summary_row, column=1).value = plot.__values['AGE']  # Age (years)
        elif 'YEAR' in PLOT_VARS:  # if not, the year of the treatment...
            ws_summary.cell(row=summary_row, column=1).value = plot.__values['YEAR']  # Year
        else:  # if not, the scenario age
            if first_line == True:
                global initial_scenario_age
                ws_summary.cell(row=summary_row, column=1).value = initial_scenario_age  # Scenario age (years)
            else:
                ws_summary.cell(row=summary_row, column=1).value = ws_summary.cell(row=summary_row - 1, column=1).value + time  # Scenario age (years)

        ws_summary.cell(row=summary_row, column=2).value = round(plot.__values['DOMINANT_H'], dec_pts) if type(plot.__values['DOMINANT_H']) == int or type(plot.__values['DOMINANT_H']) == float else plot.__values['DOMINANT_H']  # Ho (m)    

        # 
        # Masa principal antes de la clara - Stand conditions before a cut
        # 

        ws_summary.cell(row=summary_row, column=3).value = round(plot.__values['DENSITY'], dec_pts) if type(plot.__values['DENSITY']) == int or type(plot.__values['DENSITY']) == float else plot.__values['DENSITY']  # N (trees/ha)
        ws_summary.cell(row=summary_row, column=4).value = round(plot.__values['QM_DBH'], dec_pts) if type(plot.__values['QM_DBH']) == int or type(plot.__values['QM_DBH']) == float else plot.__values['QM_DBH']  # Dg (cm)

        global ba_antes
        ba_antes = plot.__values['BASAL_AREA']  # that variable is needed to calculate DG extraida for stand models

        ws_summary.cell(row=summary_row, column=5).value = round(plot.__values['BASAL_AREA'], dec_pts) if type(plot.__values['BASAL_AREA']) == int or type(plot.__values['BASAL_AREA']) == float else plot.__values['BASAL_AREA']  # G (m2/ha)                
        
        if 'VOL' in PLOT_VARS and plot.__values['VOL'] != '':
            ws_summary.cell(row=summary_row, column=6).value = round(plot.__values['VOL'], dec_pts) if type(plot.__values['VOL']) == int or type(plot.__values['VOL']) == float else plot.__values['VOL']  # V (m3/ha)
        else:
            ws_summary.cell(row=summary_row, column=6).value = '-'  # V (m3/ha)            
    

    def muerta_tree(self, plot, ws_summary, summary_row, dec_pts, labels):
        """
        Function to print the plot information of dead trees at the summary sheet of the output.
        That function will run only in individual tree models.
        """

        # 
        # Masa muerta - Dead trees
        #               

        if 'DEAD_DENSITY' in PLOT_VARS and plot.dead_density != '':  # if we have dead trees to eliminate from the plot...

            if not isinstance(ws_summary.cell(row=7, column=14).value, str):  # condition needed to print the headers only once

                # We print the headers...
                ws_summary.merge_cells('N6:P6', 6, 14)
                ws_summary.cell(row=6, column=14).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=6, column=14).value = labels['simanfor.general.stand_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=14).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=14).value = labels['simanfor.general.sum_density_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=15).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=15).value = labels['simanfor.general.sum_qmdbh_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=16).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=16).value = labels['simanfor.general.sum_vol_dead']  # Write the name of the label in the cell, in order to modify it using different languages

        # if len(plot.__dead_trees) != 0:  # if we have dead trees to eliminate from the plot...
            #n_dead = ba_dead = vol_dead = 0
            #for tree in plot.__dead_trees.values():  # tree variables calculation for dead tree information
            #    n_dead += tree.expan
            #    ba_dead += tree.expan * tree.basal_area / 10000
            #    vol_dead += tree.expan * tree.vol / 1000

            # ...and the data
            ws_summary.cell(row=summary_row, column=14).value = round(plot.dead_density, dec_pts) if type(plot.dead_density) == int or type(plot.dead_density) == float else plot.dead_density  # N (trees/ha)
           
            try:  # that condition catch the possible error of the next calculation
                if plot.dead_density != 0:
                    dg = 200 * math.pow((plot.dead_ba/math.pi/plot.dead_density), 0.5)
                else:
                    dg = 0
                ws_summary.cell(row=summary_row, column=15).value = round(dg, dec_pts) if type(dg) == int or type(dg) == float else dg  # Dg (cm)
            except ZeroDivisionError as e:
                print(e)
            
            if 'DEAD_VOL' in PLOT_VARS and plot.dead_vol == '':
                plot.add_value('DEAD_VOL', 0)
            
            if 'DEAD_VOL' in PLOT_VARS:
                ws_summary.cell(row=summary_row, column=16).value = round(plot.dead_vol, dec_pts) if type(plot.dead_vol) == int or type(plot.dead_vol) == float else plot.dead_vol  # V (m3/ha)    
            else:
                ws_summary.cell(row=summary_row, column=16).value = '-'  # V (m3/ha)                 


    def muerta_stand(self, plot, ws_summary, summary_row, dec_pts, labels):
        """
        Function to print the plot information of dead trees at the summary sheet of the output.
        That function will run only in stand models.
        """

        # 
        # Masa muerta - Dead trees
        #  

        if 'DEAD_DENSITY' in PLOT_VARS and plot.dead_density != '':  # if we have dead trees to eliminate from the plot...

            if not isinstance(ws_summary.cell(row=7, column=14).value, str):  # condition needed to print the headers only once

                # We print the headers...
                ws_summary.merge_cells('N6:P6', 6, 14)
                ws_summary.cell(row=6, column=14).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=6, column=14).value = labels['simanfor.general.stand_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=14).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=14).value = labels['simanfor.general.sum_density_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=15).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=15).value = labels['simanfor.general.sum_qmdbh_dead']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=16).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=16).value = labels['simanfor.general.sum_vol_dead']  # Write the name of the label in the cell, in order to modify it using different languages

            # ...and the data
            if isinstance(ws_summary.cell(row=(summary_row - 1), column=3).value, float):  # if it is not the first line...

                #if isinstance(ws_summary.cell(row=(summary_row - 1), column=7).value, float):  # if it was a previus harvest process...
                #    n_dead = ws_summary.cell(row=(summary_row - 1), column=3).value - ws_summary.cell(row=summary_row, column=3).value - ws_summary.cell(row=(summary_row - 1), column=7).value
                #else:
                #    n_dead = ws_summary.cell(row=(summary_row - 1), column=3).value - ws_summary.cell(row=summary_row, column=3).value
                n_dead = plot.dead_density
                ws_summary.cell(row=summary_row, column=14).value = round(n_dead, dec_pts) if type(n_dead) == int or type(n_dead) == float else n_dead  # N (trees/ha)

                if 'DEAD_BA' in PLOT_VARS and plot.dead_ba != '':
                    if n_dead != 0:
                        MTBA = plot.dead_ba*10000/n_dead  # Mean Tree Basal Area
                        dg = 2*math.sqrt(MTBA/math.pi)  # Quadratic mean diameter
                    else: 
                        dg = 0
                    ws_summary.cell(row=summary_row, column=15).value = round(dg, dec_pts) if type(dg) == int or type(dg) == float else dg  # dg (cm2)

                if 'DEAD_VOL' in PLOT_VARS and plot.dead_vol != '':
                    vol_dead = plot.dead_vol
                    ws_summary.cell(row=summary_row, column=16).value = round(vol_dead, dec_pts) if type(vol_dead) == int or type(vol_dead) == float else vol_dead  # V (m3/ha)            

    
    def extraida_tree(self, plot, ws_summary, summary_row, dec_pts):
        """
        Function to print the plot information of the extracted wood by a cut at the summary sheet of the output.
        That function will run only in individual tree models.
        """

        # 
        # Masa extraída - Cut trees
        # 

        # The next conditions are important in order to get the previous value of the variables, on the cases when consecutive cuts are planned
        if ws_summary.cell(row=summary_row, column=7).value != None:
            old_n_cut = ws_summary.cell(row=summary_row, column=7).value
        else:
            old_n_cut = 0
        if ws_summary.cell(row=summary_row, column=8).value != None:
            old_ba_cut = ws_summary.cell(row=summary_row, column=8).value
        else:
            old_ba_cut = 0
        if ws_summary.cell(row=summary_row, column=9).value != None:
            old_vol_cut = ws_summary.cell(row=summary_row, column=9).value
        else:
            old_vol_cut = 0

        # If we got some value from a past harvest, we get it to our calculation variables - that process is needed to % of cut variables calculation
        n_cut = old_n_cut
        ba_cut = old_ba_cut
        vol_cut = old_vol_cut

        for tree in plot.__cut_trees.values():  # to the previous value (int or 0), we add the new values from the cut trees of the harvest
            n_cut += tree.expan
            ba_cut += tree.expan*tree.basal_area/10000
            if 'vol' in TREE_VARS:
                if tree.vol != '':
                    vol_cut += tree.expan*tree.vol/1000

        if n_cut != 0:

            # That variables show, at the plot sheet, the % of each cut variable that is reduced by each harvest process
            if self.__values['DENSITY'] != 0 and self.__values['DENSITY'] != '':
                plot.add_value('DENSITY_CUT_VOLUME', ((n_cut - old_n_cut)/self.__values['DENSITY'])*100)
            if self.__values['BASAL_AREA'] != 0 and self.__values['BASAL_AREA'] != '':
                plot.add_value('BA_CUT_VOLUME', ((ba_cut - old_ba_cut)/self.__values['BASAL_AREA'])*100)
            if self.__values['VOL'] != 0 and self.__values['VOL'] != '':
                plot.add_value('VOL_CUT_VOLUME', ((vol_cut - old_vol_cut)/self.__values['VOL'])*100)

            ws_summary.cell(row=summary_row, column=7).value = round(n_cut, dec_pts) if type(n_cut) == int or type(n_cut) == float else n_cut  # N (trees/ha)

            try:  # that condition catch the possible error of the next calculation
                dg_cm = 200 * math.pow((ba_cut/math.pi/n_cut), 0.5)
                ws_summary.cell(row=summary_row, column=8).value = round(dg_cm, dec_pts) if type(dg_cm) == int or type(dg_cm) == float else dg_cm  # Dg (cm)
            except ZeroDivisionError as e:
                print(e)

            ws_summary.cell(row=summary_row, column=9).value = round(vol_cut, dec_pts) if type(vol_cut) == int or type(vol_cut) == float else vol_cut  # V (m3/ha)


    def extraida_plot(self, plot, ws_summary, summary_row, dec_pts):
        """
        Function to print the plot information of the extracted wood by a cut at the summary sheet of the output.
        That function will run only in stand models.        
        """

        # 
        # Masa extraída - Cut trees
        # 

        self.despues(plot, ws_summary, summary_row, dec_pts)  # it's needed to run that before, in order to use that data at the next processes

        # N - calculated by differences into the values after and before the harvest
        n_cut = ws_summary.cell(row=summary_row, column=3).value - ws_summary.cell(row=summary_row, column=10).value
        ws_summary.cell(row=summary_row, column=7).value = round(n_cut, dec_pts) if type(n_cut) == int or type(n_cut) == float else n_cut  # N (trees/ha)

        # DG - calculated by differences into the values after and before the harvest
        global ba_antes
        global ba_despues
        ba_cut = ba_antes - ba_despues  # using global variables created in other processes, we calculate the basal_area cut by the harvest

        try:  # that condition catch the possible error of the next calculation
            if n_cut != 0:
                dg_cm = 200*math.pow((ba_cut/math.pi/n_cut), 0.5)
            else:
                dg_cm = 0
            ws_summary.cell(row=summary_row, column=8).value = round(dg_cm, dec_pts) if type(dg_cm) == int or type(dg_cm) == float else dg_cm  # Dg (cm)    
        except ZeroDivisionError as e:
            print(e)

        # VOL - calculated by differences into the values after and before the harvest
        vol_cut = ws_summary.cell(row=summary_row, column=6).value - ws_summary.cell(row=summary_row, column=13).value
        ws_summary.cell(row=summary_row, column=9).value = round(vol_cut, dec_pts) if type(vol_cut) == int or type(vol_cut) == float else vol_cut  # V (m3/ha)


    def despues(self, plot, ws_summary, summary_row, dec_pts):
        """
        Function to print the plot information after a cut at the summary sheet of the output.
        """

        # 
        # Masa principal después de la clara - Stand after cut
        #                         

        ws_summary.cell(row=summary_row, column=10).value = round(plot.__values['DENSITY'], dec_pts) if type(plot.__values['DENSITY']) == int or type(plot.__values['DENSITY']) == float else plot.__values['DENSITY']  # N (trees/ha)
        ws_summary.cell(row=summary_row, column=11).value = round(plot.__values['QM_DBH'], dec_pts) if type(plot.__values['QM_DBH']) == int or type(plot.__values['QM_DBH']) == float else plot.__values['QM_DBH']  # Dg (cm)

        global ba_despues
        ba_despues = plot.__values['BASAL_AREA']  # that variable is needed to calculate DG extraida for stand models

        ws_summary.cell(row=summary_row, column=12).value = round(plot.__values['BASAL_AREA'], dec_pts) if type(plot.__values['BASAL_AREA']) == int or type(plot.__values['BASAL_AREA']) == float else plot.__values['BASAL_AREA']  # G (m2/ha)
        
        if 'VOL' in PLOT_VARS and plot.__values['VOL'] != '':
            ws_summary.cell(row=summary_row, column=13).value = round(plot.__values['VOL'], dec_pts) if type(plot.__values['VOL']) == int or type(plot.__values['VOL']) == float else plot.__values['VOL']  # V (m3/ha)
        else:
            ws_summary.cell(row=summary_row, column=13).value = '-'  # V (m3/ha)

        
    def incorporada(self, plot, ws_summary, summary_row, dec_pts, labels):
        """
        Function to print the plot information of the ingrowth at the summary sheet of the output.
        By the moment, that function is not programmed to stand models.
        """

        # 
        # Masa incorporada - Ingrowth
        #      

        if 'ING_DENSITY' in PLOT_VARS and plot.ing_density != '':  # if we have trees to add at the plot...
        # if len(plot.__ingrowth_trees) != 0:  # if we have trees to add at the plot...
        
            ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ' 
            new_columns = 0
            if 'DEAD_DENSITY' in PLOT_VARS and plot.dead_density != '':  # if dead trees exist, they are printed at the summary sheet and we need to move ingrowth data
                new_columns = 3

            if not isinstance(ws_summary.cell(row=7, column=14 + new_columns).value, str):  # condition needed to print the headers only once

                # We print the headers...        
                ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:' + str(ascii_uppercase[14 + new_columns]) + '6')

                ws_summary.cell(row=6, column=14 + new_columns).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=6, column=14 + new_columns).value = labels['simanfor.general.stand_ingrowth']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=14 + new_columns).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=14 + new_columns).value = labels['simanfor.general.sum_density_ing']  # Write the name of the label in the cell, in order to modify it using different languages

                ws_summary.cell(row=7, column=15 + new_columns).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=7, column=15 + new_columns).value = labels['simanfor.general.sum_ba_ing']  # Write the name of the label in the cell, in order to modify it using different languages

            # ...and the data
            if isinstance(plot.ing_density, str) or isinstance(plot.ing_ba, str):
                ws_summary.cell(row=summary_row, column=14 + new_columns).value = plot.ing_density  # N (trees/ha)
                ws_summary.cell(row=summary_row, column=15 + new_columns).value = plot.ing_ba  # G m2/ha                  
            else:
                ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.ing_density, dec_pts) if type(plot.ing_density) == int or type(plot.ing_density) == float else plot.ing_density  # N (trees/ha)
                ws_summary.cell(row=summary_row, column=15 + new_columns).value = round(plot.ing_ba, dec_pts) if type(plot.ing_ba) == int or type(plot.ing_ba) == float else plot.ing_ba  # G m2/ha  


    def special_data(self, plot, ws_summary, summary_row, dec_pts, labels):
        """
        Function that includes at the summary sheet the information about no timber products, if they exist in the used model.
        By the moment, the no timber products available at SIMANFOR are:
            - Cork --> Quercus suber
            - Cones and seeds --> Pinus pinea
            - Mushrooms --> Pinus sylvestris, Pinus pinaster and Pinus nigra
        """

        new_columns = 0  # Variable that will be increased on each new column added to the summary sheet, only if it is needed
        ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ' 

        if 'DEAD_DENSITY' in PLOT_VARS and plot.dead_density != '':  # if dead trees exist, they are printed at the summary sheet and we need to move ingrowth data
            new_columns += 3
        if 'ING_DENSITY' in PLOT_VARS and plot.ing_density != '':  # if dead trees exist, they are printed at the summary sheet and we need to move ingrowth data
            new_columns += 2

        if 'W_CORK' in PLOT_VARS:
        
            if not isinstance(ws_summary.cell(row=7, column=14 + new_columns).value, str):  # condition needed to print the headers only once

                QSUBER_VARS = []  # new list to include Quercus suber variables created
        
                # Now, we are using always the same structure to create the information we want to add
                QSUBER_VARS.append('QSUBER_VARS')  # Append the name of the variable, already created at the language file
                ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:' + str(ascii_uppercase[16 + new_columns]) + '6') 
                ws_summary.cell(row=6, column=14 + new_columns).font = Font(bold= True)  # Edit the cell to show the name un bold
                ws_summary.cell(row=6, column=14 + new_columns).value = labels['simanfor.general.QSUBER_VARS']  # Write the name of the label in the cell, in order to modify it using different languages

                QSUBER_VARS.append('W_CORK')
                ws_summary.cell(row=7, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=14 + new_columns).value = labels['simanfor.general.W_CORK']
    
                QSUBER_VARS.append('TOTAL_W_DEBARK')
                ws_summary.cell(row=7, column=15 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=15 + new_columns).value = labels['simanfor.general.TOTAL_W_DEBARK']

                QSUBER_VARS.append('BARK_VOL')
                ws_summary.cell(row=7, column=16 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=16 + new_columns).value = labels['simanfor.general.BARK_VOL']

                QSUBER_VARS.append('TOTAL_V_DEBARK')
                ws_summary.cell(row=7, column=17 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=17 + new_columns).value = labels['simanfor.general.TOTAL_V_DEBARK']


            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.w_cork, dec_pts) if type(plot.w_cork) == int or type(plot.w_cork) == float else plot.w_cork  # Give to the correspondent cell (summary_row) the value
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text
            new_columns += 1

            if plot.total_w_debark != '':  # that variable must be written at summary_row -1 to correct the place of the data after debark
                ws_summary.cell(row=summary_row - 1, column=14 + new_columns).value = round(plot.total_w_debark, dec_pts) if type(plot.total_w_debark) == int or type(plot.total_w_debark) == float else plot.total_w_debark  # Give to the correspondent cell (summary_row) the value
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text
            new_columns += 1

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.bark_vol, dec_pts) if type(plot.bark_vol) == int or type(plot.bark_vol) == float else plot.bark_vol  # Give to the correspondent cell (summary_row) the value
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text
            new_columns += 1

            if plot.total_v_debark != '':  # that variable must be written at summary_row -1 to correct the place of the data after debark
                ws_summary.cell(row=summary_row - 1, column=14 + new_columns).value = round(plot.total_v_debark, dec_pts) if type(plot.total_v_debark) == int or type(plot.total_v_debark) == float else plot.total_v_debark  # Give to the correspondent cell (summary_row) the value
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text
            new_columns += 1

        if 'ALL_CONES' in PLOT_VARS:

            if not isinstance(ws_summary.cell(row=7, column=14 + new_columns).value, str):  # condition needed to print the headers only once

                PPINEA_VARS = []  # new list to include Pinus pinea variables created
    
                PPINEA_VARS.append('PPINEA_VARS')
                ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:' + str(ascii_uppercase[17 + new_columns]) + '6')          
                ws_summary.cell(row=6, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=6, column=14 + new_columns).value = labels['simanfor.general.PPINEA_VARS']

                PPINEA_VARS.append('ALL_CONES')
                ws_summary.cell(row=7, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=14 + new_columns).value = labels['simanfor.general.ALL_CONES']
            
                PPINEA_VARS.append('SOUND_CONES')
                ws_summary.cell(row=7, column=15 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=15 + new_columns).value = labels['simanfor.general.SOUND_CONES']

                PPINEA_VARS.append('SOUND_SEEDS')
                ws_summary.cell(row=7, column=16 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=16 + new_columns).value = labels['simanfor.general.SOUND_SEEDS']

                PPINEA_VARS.append('W_SOUND_CONES')
                ws_summary.cell(row=7, column=17 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=17 + new_columns).value = labels['simanfor.general.W_SOUND_CONES']

                PPINEA_VARS.append('W_ALL_CONES')
                ws_summary.cell(row=7, column=18 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=18 + new_columns).value = labels['simanfor.general.W_ALL_CONES']

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.all_cones, dec_pts) if type(plot.all_cones) == int or type(plot.all_cones) == float else plot.all_cones
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text         
            new_columns += 1

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.sound_cones, dec_pts) if type(plot.sound_cones) == int or type(plot.sound_cones) == float else plot.sound_cones
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text
            new_columns += 1

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.sound_seeds, dec_pts) if type(plot.sound_seeds) == int or type(plot.sound_seeds) == float else plot.sound_seeds
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 28: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 28  # resize that column to show the text
            new_columns += 1

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.w_sound_cones, dec_pts) if type(plot.w_sound_cones) == int or type(plot.w_sound_cones) == float else plot.w_sound_cones
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 30: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 30  # resize that column to show the text
            new_columns += 1

            ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.w_all_cones, dec_pts) if type(plot.w_all_cones) == int or type(plot.w_all_cones) == float else plot.w_all_cones
            if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 30: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 30  # resize that column to show the text
            new_columns += 1

        if 'EDIBLE_MUSH' in PLOT_VARS:

            if not isinstance(ws_summary.cell(row=7, column=14 + new_columns).value, str):  # condition needed to print the headers only once

                MUSHROOMS_VARS = []  # new list to include mushroom variables created
            
                MUSHROOMS_VARS.append('MUSHROOMS_VARS')       
                if len(ascii_uppercase) > (13 + new_columns) and len(ascii_uppercase) > (15 + new_columns):
                    ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:' + str(ascii_uppercase[15 + new_columns]) + '6')   
                elif len(ascii_uppercase) > (13 + new_columns) and len(ascii_uppercase) <= (15 + new_columns):
                    ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:A' + str(ascii_uppercase[15 + new_columns - 26]) + '6')                                       
                else:
                    ws_summary.merge_cells(str('A' + ascii_uppercase[13 + new_columns]) + '6:A' + str(ascii_uppercase[15 + new_columns - 26]) + '6')                      
                ws_summary.cell(row=6, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=6, column=14 + new_columns).value = labels['simanfor.general.MUSHROOMS_VARS']

                MUSHROOMS_VARS.append('EDIBLE_MUSH')
                ws_summary.cell(row=7, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=14 + new_columns).value = labels['simanfor.general.EDIBLE_MUSH']

                MUSHROOMS_VARS.append('MARKETED_MUSH')
                ws_summary.cell(row=7, column=15 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=15 + new_columns).value = labels['simanfor.general.MARKETED_MUSH']

                if 'MARKETED_LACTARIUS' in PLOT_VARS:
                    MUSHROOMS_VARS.append('MARKETED_LACTARIUS')
                    ws_summary.cell(row=7, column=16 + new_columns).font = Font(bold= True)
                    ws_summary.cell(row=7, column=16 + new_columns).value = labels['simanfor.general.MARKETED_LACTARIUS']

            if plot.edible_mush != '':  # print it just if it is calculated
                ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.edible_mush, dec_pts) if type(plot.edible_mush) == int or type(plot.edible_mush) == float else plot.edible_mush
    
            if len(ascii_uppercase) > (13 + new_columns):
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text   
            else:
                if ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width < 25: ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width = 25  # resize that column to show the text
            new_columns += 1

            if plot.marketed_mush != '':  # print it just if it is calculated
                ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.marketed_mush, dec_pts) if type(plot.marketed_mush) == int or type(plot.marketed_mush) == float  else plot.marketed_mush
            
            if len(ascii_uppercase) > (13 + new_columns): 
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 28: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 28  # resize that column to show the text   
            else:
                if ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width < 28: ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width = 28  # resize that column to show the text
            new_columns += 1

            if 'MARKETED_LACTARIUS' in PLOT_VARS:  # some models have other mushrooms and not lactarius        
                if plot.marketed_lactarius != '':  # print it just if it is calculated
                    ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.marketed_lactarius, dec_pts) if type(plot.marketed_lactarius) == int or type(plot.marketed_lactarius) == float  else plot.marketed_lactarius
    
            if len(ascii_uppercase) > (13 + new_columns):
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 30: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 30  # resize that column to show the text   
            else:
                if ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width < 30: ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width = 30  # resize that column to show the text
            new_columns += 1

        if 'MUSHROOM_PRODUCTIVITY' in PLOT_VARS:

            if not isinstance(ws_summary.cell(row=7, column=14 + new_columns).value, str):  # condition needed to print the headers only once

                MUSHROOMS_VARS = []  # new list to include mushroom variables created
            
                MUSHROOMS_VARS.append('MUSHROOM_PRODUCTIVITY')       
                if len(ascii_uppercase) > (13 + new_columns) and len(ascii_uppercase) > (13 + new_columns):
                    ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:' + str(ascii_uppercase[13 + new_columns]) + '6')   
                elif len(ascii_uppercase) > (13 + new_columns) and len(ascii_uppercase) <= (13 + new_columns):
                    ws_summary.merge_cells(str(ascii_uppercase[13 + new_columns]) + '6:A' + str(ascii_uppercase[13 + new_columns - 26]) + '6')                                       
                else:
                    ws_summary.merge_cells(str('A' + ascii_uppercase[13 + new_columns]) + '6:A' + str(ascii_uppercase[13 + new_columns - 26]) + '6')                      
                ws_summary.cell(row=6, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=6, column=14 + new_columns).value = labels['simanfor.general.MUSHROOMS_VARS']

                MUSHROOMS_VARS.append('MUSHROOM_PRODUCTIVITY')
                ws_summary.cell(row=7, column=14 + new_columns).font = Font(bold= True)
                ws_summary.cell(row=7, column=14 + new_columns).value = labels['simanfor.general.MUSHROOM_PRODUCTIVITY']
            
            if plot.mushroom_productivity != '':  # print it just if it is calculated
                ws_summary.cell(row=summary_row, column=14 + new_columns).value = round(plot.mushroom_productivity, dec_pts) if type(plot.mushroom_productivity) == int or type(plot.mushroom_productivity) == float else plot.mushroom_productivity
            
            if len(ascii_uppercase) > (13 + new_columns):
                if ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width < 25: ws_summary.column_dimensions[str(ascii_uppercase[13 + new_columns])].width = 25  # resize that column to show the text   
            else:
                if ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width < 25: ws_summary.column_dimensions[str('A' + ascii_uppercase[13 + new_columns - 26])].width = 25  # resize that column to show the text
            new_columns += 1
 

    def plot_to_xslt(self, labels, workbook, row: int, next_plot, next_operation, operation, 
                       summary_row: int, dec_pts: int = 2):
        """
        Function to set how to print summary and plot sheets at the output xlsx file.
        It has a different way to print the summary sheets depending on the model type (tree or stand).
        """

        # 
        # Resumen - Summary
        # 
  
        # IT'S IMPORTANT TO BE CAREFUL WITH THE NEXT THINGS:
        #   - DEPENDING ON YOUR PURPOSE, THE INFORMATION TO SEND TO THE FUNCTIONS CAN BE self (PRESENT) OR next_plot (FUTURE)
        #   - IT HAS A DIFFERENT STRUCTURE AND FUNCTIONS DEPENDING IF YOU ARE GOING TO WORK WITH TREE OR STAND MODELS
        #   - THAT SHEET HAS ALL THE INFORMATION CONDENSED (summary), SO THE INFORMATION TO SHOW MUST HAVE A GOOD STRUCTURE

        ws_summary = workbook[labels['simanfor.general.summary_sheet']]   
        
        operation_code = operation.type.get_code_name()    
        
        min = next_operation.get_variable('min_age') if (next_operation != None and next_operation.has('min_age')) else 0  # import mininum age value from scenario file
        max = next_operation.get_variable('max_age') if (next_operation != None and next_operation.has('max_age')) else 1000  # import maximum age value from scenario file

        if 'AGE' in PLOT_VARS:
            plot_age = self.age
        else:
            plot_age = 0
        
        # that 4 lines are created to show the scenario age at the summary sheet only if stand age or year are not availble
        if operation_code == 'LOAD':
            global initial_scenario_age
            initial_scenario_age = operation.get_variable('init')
        time = operation.get_variable('time')
        
                        ####################----- STAND MODELS -----####################


        if operation_code != 'LOAD' and 'stand' in Model.model_name:  # if the used model is a stand model...

            if min <= plot_age <= max:  # if the plot is not on the age range stablished on the scenario, that information will not be printed

                if operation_code == 'INIT':  # Initialization proccess

                    if next_operation != None and next_operation.type.get_code_name() == 'EXECUTION':
                    # if it is followed by an EXECUTION, only the "before cut" column is needed on this row
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                         
        
                    elif next_operation != None and next_operation.type.get_code_name() == 'HARVEST':
                    # if it is followed by an HARVEST, columns with that information are added at the same row
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.extraida_plot(next_plot, ws_summary, summary_row, dec_pts)
                        self.despues(next_plot, ws_summary, summary_row, dec_pts)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                         
        
                    elif next_operation == None:
                    # if no more operations are planned, only the "before cut" column is needed on this row                          
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                         
                        
                    summary_row += 1  # after print a row, we go to the next

                elif operation_code == 'EXECUTION':   # Execution proccess
        
                    if next_operation != None and next_operation.type.get_code_name() == 'EXECUTION':
                    # if it is followed by another EXECUTION, that row will contain only the EXECUTION information                        
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.muerta_stand(next_plot, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(next_plot, ws_summary, summary_row, dec_pts, labels)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels) 

                    elif next_operation != None and next_operation.type.get_code_name() == 'HARVEST':
                    # if it is followed by an HARVEST, the cut information (FUTURE --> next_plot) will be added to that row (PRESENT --> self)                          
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.extraida_plot(next_plot, ws_summary, summary_row, dec_pts)
                        self.despues(next_plot, ws_summary, summary_row, dec_pts)
                        self.muerta_stand(self, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(self, ws_summary, summary_row, dec_pts, labels)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels) 

                        #summary_row -= 1  # IMPORTANT! We need to use the same row to the HARVEST proccess, if not, calculations will not be succesfull

                    elif next_operation == None:
                    # if it is the last proccess, that row will contain only the EXECUTION information                        
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.muerta_stand(self, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(self, ws_summary, summary_row, dec_pts, labels)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels) 

                    summary_row += 1  # after print a row, we go to the next

                elif operation_code == 'HARVEST' and next_operation != None and next_operation.type.get_code_name() == 'HARVEST':  # Consecutive harvests
                # if a HARVEST is followed by another HARVEST, we reutilise the data of the last proccess without print it, showing consecutive cuts as only 1 proccess
                    
                    summary_row -= 1  # come back to the line before in order to overwrite the harvest information, adding the new one

                    self.extraida_plot(next_plot, ws_summary, summary_row, dec_pts)
                    self.despues(next_plot, ws_summary, summary_row, dec_pts)

                    summary_row += 1  # after print a row, we go to the next


                        ####################----- TREE MODELS -----####################


        elif operation_code != 'LOAD' and 'stand' not in Model.model_name:  # if the used model is a individual tree model...
            
            if min <= plot_age <= max:  # if the plot is not on the age range stablished on the scenario, that information will not be printed

                if operation_code == 'INIT':  # Initialization proccess

                    if next_operation != None and next_operation.type.get_code_name() == 'EXECUTION':
                    # if it is followed by an EXECUTION, only the "before cut" column is needed on this row
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)

                    elif next_operation != None and next_operation.type.get_code_name() == 'HARVEST':
                    # if it is followed by an HARVEST, columns with that information are added at the same row
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.extraida_tree(next_plot, ws_summary, summary_row, dec_pts)
                        self.despues(next_plot, ws_summary, summary_row, dec_pts)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                        

                    elif next_operation == None:
                    # if no more operations are planned, only the "before cut" column is needed on this row                        
                        self.antes(self, ws_summary, summary_row, dec_pts, time, True)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)

                    summary_row += 1  # after print a row, we go to the next

                elif operation_code == 'EXECUTION':  # Execution proccess
        
                    if next_operation != None and next_operation.type.get_code_name() == 'EXECUTION':
                    # if it is followed by another EXECUTION, that row will contain only the EXECUTION information
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.muerta_tree(self, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(self, ws_summary, summary_row, dec_pts, labels)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                        
        
                    elif next_operation != None and next_operation.type.get_code_name() == 'HARVEST':
                    # if it is followed by an HARVEST, the cut information (FUTURE --> next_plot) will be added to that row (PRESENT --> self)  
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.muerta_tree(self, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(self, ws_summary, summary_row, dec_pts, labels)
                        self.extraida_tree(next_plot, ws_summary, summary_row, dec_pts)
                        self.despues(next_plot, ws_summary, summary_row, dec_pts)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                        
                        # next line is not correct, but I leave it as a comment just in case I need to change it
                        #summary_row -= 1  # IMPORTANT! We need to use the same row to the HARVEST proccess, if not, calculations will not be succesfull

                    elif next_operation == None:
                    # if it is the last proccess, that row will contain only the EXECUTION information
                        self.antes(self, ws_summary, summary_row, dec_pts, time)
                        self.muerta_tree(self, ws_summary, summary_row, dec_pts, labels)
                        self.incorporada(self, ws_summary, summary_row, dec_pts, labels)
                        self.special_data(self, ws_summary, summary_row, dec_pts, labels)                        
                        
                    summary_row += 1  # after print a row, we go to the next

                elif operation_code == 'HARVEST' and next_operation != None and next_operation.type.get_code_name() == 'HARVEST':  # Consecutive harvests
                # if a HARVEST is followed by another HARVEST, we reutilise the data of the last proccess without print it, showing consecutive cuts as only 1 proccess
                    
                    summary_row -= 1  # come back to the line before in order to overwrite the harvest information, adding the new one

                    self.extraida_tree(next_plot, ws_summary, summary_row, dec_pts)
                    self.despues(next_plot, ws_summary, summary_row, dec_pts)

                    summary_row += 1  # after print a row, we go to the next


        self.edit_sheet(ws_summary, 'summary')  # this line is needed to center the information of the cells
        #if ws_summary.column_dimensions["J"].width > 15: ws_summary.column_dimensions["J"].width = 15  # resize that column to not make it very big


        #
        # Description sheet
        #
        ws_description = workbook[labels['simanfor.general.description_sheet']]

        # Information - Plot information - titles
        if 'REINEKE_VALUE' in PLOT_VARS:
            self.merge(ws_description, 'M7:T7', 7, 13, str(self.print_value('REINEKE_VALUE', dec_pts = 3))) 
        if 'REF_SI_AGE' in PLOT_VARS:
            self.merge(ws_description, 'M8:T8', 8, 13, str(self.print_value('REF_SI_AGE', dec_pts=dec_pts))) 
        if 'SI' in PLOT_VARS:
            self.merge(ws_description, 'M9:T9', 9, 13, str(self.print_value('SI', dec_pts=dec_pts))) 
        if 'HEGYI_RADIUS' in PLOT_VARS:
            self.merge(ws_description, 'M10:T10', 10, 13, str(self.print_value('HEGYI_RADIUS', dec_pts=dec_pts))) 

        # 
        # Parcelas - plots
        # 
        ws_plot = workbook[labels['simanfor.general.plot_sheet']]

        for i in range(len(PLOT_VARS)):
            if PLOT_VARS[i] not in PLOT_VARS_NOT_PRINT:
                ws_plot.cell(row=row+1, 
                column=i+1 + len(SCENARIO_VARS)).value = self.print_value(PLOT_VARS[i], 
                    dec_pts=dec_pts)

        # Print SPECIE_ERROR on summary and plot sheets, only if it is notified by the model 
        if Warnings.specie_error == 1 or Warnings.specie_error_trees == 1:   

            for k in range(1, (len(PLOT_VARS) + len(SCENARIO_VARS)) + 1):
                ws_plot.cell(row=row+1, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+2, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+3, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+4, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            for k in range(1, 19):
                ws_summary.cell(row=4, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_summary.cell(row=5, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            if next_operation == None:
                if Warnings.specie_error == 1:
                    ws_plot.merge_cells('A' + str(row + 4) + ':Z' + str(row + 4))
                    ws_plot.cell(row=row+4, column=1).value = labels['simanfor.metadata.SPECIE_ERROR']
                    ws_summary.merge_cells('A5:R5')           
                    ws_summary.cell(row=5, column=1).value = labels['simanfor.metadata.SPECIE_ERROR']
                else:
                    ws_plot.merge_cells('A' + str(row + 4) + ':Z' + str(row + 4))
                    ws_plot.cell(row=row+4, column=1).value = labels['simanfor.metadata.SPECIE_ERROR_TREES']
                    ws_summary.merge_cells('A5:R5')           
                    ws_summary.cell(row=5, column=1).value = labels['simanfor.metadata.SPECIE_ERROR_TREES']

        # Print EXEC_ERROR on summary and plot sheets, only if it is notified by the model
        if Warnings.exec_error == 1:     
            for k in range(1, (len(PLOT_VARS) + len(SCENARIO_VARS)) + 1):
                ws_plot.cell(row=row+1, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+2, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            for k in range(1, 19):
                ws_summary.cell(row=4, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_summary.cell(row=5, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            if next_operation == None:
                ws_plot.merge_cells('A' + str(row + 2) + ':Z' + str(row + 2))
                ws_plot.cell(row=row+2, column=1).value = labels['simanfor.metadata.EXEC_ERROR']
                ws_summary.merge_cells('A5:R5')           
                ws_summary.cell(row=5, column=1).value = labels['simanfor.metadata.EXEC_ERROR_SUMMARY']

        # Print CUT_ERROR on summary and plot sheets, only if it is notified by the model
        if Warnings.cut_error == 1: 
            for k in range(1, (len(PLOT_VARS) + len(SCENARIO_VARS)) + 1):
                ws_plot.cell(row=row+1, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+2, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_plot.cell(row=row+3, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            for k in range(1, 19):
                ws_summary.cell(row=4, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
                ws_summary.cell(row=5, column=k).border = Border(bottom = Side(border_style = 'thick', color = 'FF0F0F'))
            if next_operation == None:
                ws_plot.merge_cells('A' + str(row+3) + ':Z' + str(row+3)) 
                ws_plot.cell(row=row+3, column=1).value = labels['simanfor.metadata.CUT_ERROR']
                ws_summary.merge_cells('A5:R5')
                ws_summary.cell(row=5, column=1).value = labels['simanfor.metadata.CUT_ERROR_SUMMARY']

        self.edit_sheet(ws_plot, 'plot')  # this line is needed to edit the page
        if ws_plot.column_dimensions["A"].width > 15: ws_plot.column_dimensions["A"].width = 15  # resize that column to not make it very big

        return summary_row


    def trees_to_xlst(self, labels, workbook, node, print_trees=False, decimals: int = 2):
        """
        Function to set how to print tree sheets information at the output xlsx file.
        *all_in_one* command can print all the trees in the same sheet.
        """

        # Code to not print trees sheets at the ouput of stand models
        stand = str('stand')
        if stand in Model.model_name:
            print_trees = False

        # variable that write all the nodes at the same sheet
        # that will write all the trees in a sheet called 'Trees'
        all_in_one = False  # set True to print all the trees in the same sheet

        row = 1
        node = node - 1  # change needed to start into node nº 1 (initialization), and set node 0 as data import

        if node == 0:  # at the first node (initial inventory), we use another method to show all the information, not deleting tree variables

            ws_node = workbook.create_sheet(labels['simanfor.general.initial_inventory'])

            ws = workbook.active

            # Edit the color of the selected sheet
            ws_node.sheet_properties.tabColor = "03B300"
            
            for i in range(len(Tree.variables_names_original())):  # print the name of variables (first line on the output)
                ws_node.cell(row=row, column=i+1).value = labels['simanfor.tree.' + Tree.variables_names_original()[i]]

            tree_n = 0  # counter of trees, needed to select the line to fill with red color

            for tree in self.__trees.values():  # print alive trees
                # at node 0 (INIT), we just want to copy the initial inventory, not marking trees from another species
                row += 1
                tree_n += 1
                tree.to_xslt_original(ws_node, row, decimals)

        elif node != 0 and print_trees == True and all_in_one == False:  

            # following nodes are written only with the information (variables) desired by the model
            # they are written only if the model type is tree; to stand models, as tree data doesn't change, they aren't needed

            ws_node = workbook.create_sheet(labels['simanfor.general.node_sheet'] + ' ' + str(node) + ' - ' + labels['simanfor.general.trees_sheet'])

            ws = workbook.active

            # Edit the color of the selected sheet
            ws_node.sheet_properties.tabColor = "03B300"

            for i in range(len(Tree.variables_names())):  # print the name of variables (first line on the output)
                ws_node.cell(row=row, column=i+1).value = labels['simanfor.tree.' + Tree.variables_names()[i]]

            tree_n = 0  # counter of trees, needed to select the line to fill with red color

            order_criteria = OrderCriteria(DESC)  # we stablish the criteria to order the trees at the output
            order_criteria.add_criteria('TREE_ID')  # we stablish the order of the output trees by id

            alive_trees: Tree = Tree.get_sord_and_order_tree_list(self.__trees.values(), order_criteria=order_criteria)  
            for tree in alive_trees:  # print alive trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)


            dead_trees: Tree = Tree.get_sord_and_order_tree_list(self.__dead_trees.values(), order_criteria=order_criteria)
            for tree in dead_trees:  # print dead trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            cut_trees: Tree = Tree.get_sord_and_order_tree_list(self.__cut_trees.values(), order_criteria=order_criteria)
            for tree in cut_trees:  # print cut trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            ingrowth_trees: Tree = Tree.get_sord_and_order_tree_list(self.__ingrowth_trees.values(), order_criteria=order_criteria)
            for tree in ingrowth_trees:  # print added trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            last_col = len(Tree.variables_names())  # locate "status" position
            ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
            ws_node.insert_cols(4)  # insert a new column to copy status information in there

            if last_col < len(ascii_uppercase):  # select the letter of the cell we want to move
                cell_to_move =  ascii_uppercase[last_col]
            elif last_col >= len(ascii_uppercase) and last_col < len(ascii_uppercase)*2:
                cell_to_move =  ascii_uppercase[0] + ascii_uppercase[last_col - len(ascii_uppercase)]
            elif last_col >= len(ascii_uppercase)*2 and last_col < len(ascii_uppercase)*3:
                cell_to_move =  ascii_uppercase[1] + ascii_uppercase[last_col - len(ascii_uppercase)*2]

            # move the desired column to their new position
            ws_node.move_range(cell_to_move + str(1) + ':' + cell_to_move + str(ws_node.max_row), rows=0, cols= - len(Tree.variables_names()) + 3)

            self.edit_sheet(ws_node, 'nodes')  # edit the sheet of the output


        elif node != 0 and print_trees == True and all_in_one == True:  

            # following nodes are written only with the information (variables) desired by the model
            # they are written only if the model type is tree; to stand models, as tree data doesn't change, they aren't needed

            if node == 1:  # create the sheet only once
                ws_node = workbook.create_sheet('Trees')

            ws_node = workbook['Trees']   

            ws = workbook.active

            # Edit the color of the selected sheet
            ws_node.sheet_properties.tabColor = "03B300"

            # get the row number of the last node
            global last_row_on_tree_sheet
            row = last_row_on_tree_sheet

            if node == 1:
                for i in range(len(Tree.variables_names())):  # print the name of variables (first line on the output)
                    ws_node.cell(row=row, column=i+1).value = labels['simanfor.tree.' + Tree.variables_names()[i]]

            tree_n = 0  # counter of trees, needed to select the line to fill with red color

            order_criteria = OrderCriteria(DESC)  # we stablish the criteria to order the trees at the output
            order_criteria.add_criteria('TREE_ID')  # we stablish the order of the output trees by id

            alive_trees: Tree = Tree.get_sord_and_order_tree_list(self.__trees.values(), order_criteria=order_criteria)  
            for tree in alive_trees:  # print alive trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)


            dead_trees: Tree = Tree.get_sord_and_order_tree_list(self.__dead_trees.values(), order_criteria=order_criteria)
            for tree in dead_trees:  # print dead trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            cut_trees: Tree = Tree.get_sord_and_order_tree_list(self.__cut_trees.values(), order_criteria=order_criteria)
            for tree in cut_trees:  # print cut trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            ingrowth_trees: Tree = Tree.get_sord_and_order_tree_list(self.__ingrowth_trees.values(), order_criteria=order_criteria)
            for tree in ingrowth_trees:  # print added trees
                row += 1
                tree_n += 1
                if Model.specie_ifn_id != 0 and Model.specie_ifn_id != '':
                    if tree.specie != Model.specie_ifn_id:
                        self.cell_color(ws_node, tree_n + 1)
                tree.to_xslt(ws_node, row, decimals)

            # update row number to the next node
            last_row_on_tree_sheet = row


    def cell_color(self, sheet, tree_n):
        """
        Function that fill the selected cells in a color as a warning message.
        That function is called from trees_to_xlst in order to mark the trees of a different specie that the used model.
        """

        redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # set red color
        
        n = 0  # lines counter
        z = 2  # number to multiply the lenght of the list
        ascii_uppercase='ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        for k in Tree.variables_names():
            if n < len(ascii_uppercase):
                sheet[ascii_uppercase[n] + str(tree_n)].fill = redFill
                n += 1

            else:  # once alphabet is finished, we add another first letter
                if n < len(ascii_uppercase)*z:
                    sheet[ascii_uppercase[z - 2] + ascii_uppercase[n - len(ascii_uppercase)*(z - 1)] + str(tree_n)].fill = redFill
                    n += 1

                else:
                    z += 1
                    sheet[ascii_uppercase[z - 2] + ascii_uppercase[n - len(ascii_uppercase)*(z - 1)] + str(tree_n)].fill = redFill
                    n += 1

    def edit_sheet(self, sheet, page):
        """
        Function that center, bold titles and auto-size all the cells of a page.
        """

        if page != 'summary':  # code to edit the pages

            # cell auto-size
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                if length < 7:
                    length = 6
                sheet.column_dimensions[column_cells[0].column_letter].width = length + 1

            # center all cells
            for col in sheet.columns:
                count_lines = 0
                for cell in col:
                    # openpyxl styles aren't mutable,
                    # so you have to create a copy of the style, modify the copy, then set it back

                    if count_lines == 0:  # condition to center headers but not the rest of lines
                        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')                        
                        count_lines += 1
                    else:      
                        alignment_obj = cell.alignment.copy(horizontal='left', vertical='center')

                    cell.alignment = alignment_obj

            # set bold titles of the columns
            for cell in sheet["1:1"]:
                cell.font = Font(bold=True)

        else:  # Summary sheet only need to center the contents, the rest is done on simulation.py
            cols_not_center = ['F', 'G', 'H', 'K', 'L', 'M']
            # center all cells
            for col in sheet.columns:
                count_lines = 1                
                
                for cell in col:
                    # openpyxl styles aren't mutable,
                    # so you have to create a copy of the style, modify the copy, then set it back
                    
                    # the following conditions are only needed just to not set at left the joined cells
                    if count_lines < 8:
                        alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                        count_lines += 1
                    else:
                        alignment_obj = cell.alignment.copy(horizontal='left', vertical='center')

                    cell.alignment = alignment_obj


    def merge(self, sheet, cell_range, row, col, val):
        """
        Function used to create a merged cell at the output xlsx file.
        """

        sheet.merge_cells(cell_range)
        cell = sheet.cell(row = row, column = col)
        cell.value = val                    

global initial_scenario_age, last_row_on_tree_sheet
initial_scenario_age = 0
last_row_on_tree_sheet = 1