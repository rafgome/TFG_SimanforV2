#!/usr/bin/env python3
#
# Copyright (c) $today.year Spiros Michalakopoulos (Sngular). All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ==============================================================================
# import pdb

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font


num_diffs = 0

workbook_diffs = 0
sheet_diffs = list()
cell_value_diffs = list()

type_diffs = ['workbook', 'sheet', 'cell']


def check(val1, val2, type_diff: str, idx: int, name: str, row = None, col = None):

    try:
        assert(val1 == val2)
    except AssertionError:
        global num_diffs
        num_diffs += 1

        if type_diff == 'sheet':
            diff = "val1: " + str(val1) + ",  val2: " + str(val2)
            sheet_diffs.append((str(idx), name, diff))

        elif type_diff == 'cell':
            diff = "val1: " + str(val1) + ",  val2: " + str(val2)
            cell_value_diffs.append((str(idx), name, row, col, diff))

    return 

   
def test_sheet(wb1, wb2, idx: int, name: str):
    
    global num_diffs

    try:
        sheet_resumen1 = wb1.worksheets[idx]
        sheet_resumen2 = wb2.worksheets[idx]
    except Exception as e:
        print(e)
        # pdb.set_trace()        

    check(sheet_resumen1.title, name, 'sheet', idx, name)
    check(sheet_resumen2.title, name, 'sheet', idx, name)
    check(sheet_resumen1.max_column, sheet_resumen2.max_column, 'sheet', idx, name)
    check(sheet_resumen1.max_row, sheet_resumen2.max_row, 'sheet', idx, name)

    for i in range(sheet_resumen1.min_row, sheet_resumen1.max_row):
        for j in range(sheet_resumen1.min_column, sheet_resumen1.max_column):
            
            try:
                check(sheet_resumen1.cell(i, j).value, sheet_resumen2.cell(i, j).value, 
                      'cell', idx, name, row = i, col = sheet_resumen1.cell(i, j).column_letter)
            except AttributeError:
                # *** AttributeError: 'MergedCell' object has no attribute 'column_letter'
                continue

            if num_diffs >= 100:
                return


def test_resumen(wb1, wb2):
    test_sheet(wb1, wb2, 0, 'Resumen')
    
        
def test_resumen_nodos(wb1, wb2):
    test_sheet(wb1, wb2, 1, 'Resumen nodos')


def test_parcelas(wb1, wb2):
    test_sheet(wb1, wb2, 2, 'Parcelas')

def italic(cell, label: str):
    cell.value = label
    cell.font = Font(name = 'Calibri', italic = True)            

def bold(cell, label: str):
    cell.value = label
    cell.font = Font(name = 'Calibri', bold = True)            

def test_wip():

    files1 = [
                r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_5 copy 3.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_2.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_3.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_4.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_5.xlsx'
                ]
    
    files2 = [
                r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_5.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_2 copy 2.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_3 copy 2.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_4 copy 2.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/output/output_release/Output_Plot_5 copy 2.xlsx'
                ]
    
    reports = [
                r'/Users/s.michalakapoulos/dev/simanfor/simanfor_cortas/simulator/tests/regtests/output/report1.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/simanfor_cortas/simulator/tests/regtests/output/report2.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/simanfor_cortas/simulator/tests/regtests/output/report3.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/simanfor_cortas/simulator/tests/regtests/output/report4.xlsx', 
                # r'/Users/s.michalakapoulos/dev/simanfor/simanfor_cortas/simulator/tests/regtests/output/report5.xlsx'
                ]
    
    assert(len(files1) == len(files2))
    assert(len(reports) == len(files2))

    for i in range(len(files1)):
        main(files1[i], files2[i], reports[i])
        

def test_scenario_claras_01():

    files1 = [
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_1.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_2.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_3.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_4.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_5.xlsx'
                ]
    
    files2 = [
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_01/Output_Plot_1.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_01/Output_Plot_2.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_01/Output_Plot_3.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_01/Output_Plot_4.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_01/Output_Plot_5.xlsx'
                ]
    
    reports = [
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_01/report1.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_01/report2.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_01/report3.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_01/report4.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_01/report5.xlsx'
                ]

    assert(len(files1) == len(files2))
    assert(len(reports) == len(files2))

    for i in range(len(files1)):
        main(files1[i], files2[i], reports[i])


def test_scenario_claras_02():

    files1 = [
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_1.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_2.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_3.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_4.xlsx',
                r'/home/spiros/dev/simanfor/output/output_release/Output_Plot_5.xlsx'
                ]
    
    files2 = [
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_02/Output_Plot_1.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_02/Output_Plot_2.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_02/Output_Plot_3.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_02/Output_Plot_4.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/input/scenario_claras_02/Output_Plot_5.xlsx'
                ]
    
    reports = [
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_02/report1.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_02/report2.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_02/report3.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_02/report4.xlsx',
                r'/home/spiros/dev/simanfor/release/simulator/tests/regtests/output/scenario_claras_02/report5.xlsx'
                ]

    assert(len(files1) == len(files2))
    assert(len(reports) == len(files2))

    for i in range(len(files1)):
        main(files1[i], files2[i], reports[i])


def main(file1, file2, report):
    
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "test sheet"
    
    cell = worksheet.cell(row = 1, column = 1)
    cell.value = "Summary"
    cell.font = Font(name = 'Calibri', bold = True)

    cell = worksheet.cell(row = 2, column = 1)
    cell.value = "file1:"
    cell.font = Font(name = 'Calibri', italic = True)                   
    
    cell = worksheet.cell(row = 3, column = 1)
    cell.value = "file2:"
    cell.font = Font(name = 'Calibri', italic = True)                
    
    cell = worksheet.cell(row = 4, column = 1)
    cell.value = "NoOfDiffs:"
    cell.font = Font(name = 'Calibri', italic = True)                

    worksheet.cell(row = 2, column = 2).value = file1
    worksheet.cell(row = 3, column = 2).value = file2

    for i in range(0, len(wb1.worksheets)):
        test_sheet(wb1, wb2, i, wb1.sheetnames[i])

    # test_sheet(wb1, wb2, 1, "Resumen")
    # test_resumen(wb1, wb2)

    global num_diffs
    worksheet.cell(row = 4, column = 2).value = num_diffs

    rownbr = 6
    if workbook_diffs != 0:
        worksheet.cell(row = rownbr, column = 1).value = "Workbook diff:"
        worksheet.cell(row = rownbr, column = 2).value = workbook_diffs
        rownbr += 2

    if sheet_diffs != 0:
        cell = worksheet.cell(row = rownbr, column = 1)
        bold(cell, "Sheet Diffs") 
        rownbr += 1
        
        cell = worksheet.cell(row = rownbr, column = 1)
        italic(cell, "Sheet ID")
        
        cell = worksheet.cell(row = rownbr, column = 2)
        italic(cell, "Sheet Name")
 
        cell = worksheet.cell(row = rownbr, column = 3)
        italic(cell, "diff")

        rownbr += 1        

        for i in range(0, len(sheet_diffs)):
            worksheet.cell(row = rownbr, column = 1).value = sheet_diffs[i][0]
            worksheet.cell(row = rownbr, column = 2).value = sheet_diffs[i][1]
            worksheet.cell(row = rownbr, column = 3).value = sheet_diffs[i][2]
            rownbr += 1

    rownbr += 1
            
    if cell_value_diffs != 0:
        cell = worksheet.cell(row = rownbr, column = 1)
        bold(cell, "Cell Value Diffs"); rownbr += 1

        cell = worksheet.cell(row = rownbr, column = 1)
        italic(cell, "Sheet ID")
        
        cell = worksheet.cell(row = rownbr, column = 2)
        italic(cell, "Sheet Name")
        
        cell = worksheet.cell(row = rownbr, column = 3)
        italic(cell, "Row") 
        
        cell = worksheet.cell(row = rownbr, column = 4)
        italic(cell, "Column")
        
        cell = worksheet.cell(row = rownbr, column = 5)
        italic(cell, "diff")
         
        rownbr += 1

        for i in range(0, len(cell_value_diffs)):
            worksheet.cell(row = rownbr, column = 1).value = cell_value_diffs[i][0]
            worksheet.cell(row = rownbr, column = 2).value = cell_value_diffs[i][1]
            worksheet.cell(row = rownbr, column = 3).value = cell_value_diffs[i][2]
            worksheet.cell(row = rownbr, column = 4).value = cell_value_diffs[i][3]
            worksheet.cell(row = rownbr, column = 5).value = cell_value_diffs[i][4]
            rownbr += 1

    workbook.close()
    
    workbook.save(report)

if __name__ == "__main__":

    # test_wip()                # work in progress, manually input files
    # test_scenario_claras_01() # LIHEHEHE...
    test_scenario_claras_02() # LIEEHHEHEHEHE...
    
